from __future__ import annotations

import base64
import json
import re
import shutil
import secrets
import time
import unicodedata
from datetime import date, datetime
from html import escape
from io import BytesIO
from pathlib import Path

import pandas as pd
import streamlit as st
import streamlit.components.v1 as components
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

try:
    from st_rsuite import date_picker as rsuite_date_picker
except Exception:  # pragma: no cover - fallback when component is unavailable
    rsuite_date_picker = None


BASE_DIR = Path(__file__).parent
DATA_DIR = BASE_DIR / "data"
ASSETS_DIR = BASE_DIR / "assets"
ANEXOS_DIR = DATA_DIR / "anexos"
EMPRESAS_DATA_DIR = DATA_DIR / "empresas"
BACKUPS_DIR = DATA_DIR / "backups"
CSV_PATH = DATA_DIR / "dados_demo.csv"
XLSX_PATH = DATA_DIR / "dados_demo.xlsx"
CONFIG_PATH = BASE_DIR / "config_clientes.json"
LOGO_FULL_PATH = ASSETS_DIR / "mh_log_logo_app_512.png"
SESSOES_PATH = DATA_DIR / "sessoes_portal.json"
SESSAO_LOGIN_SEGUNDOS = 60 * 60
BACKUP_SHEET_PREFIX = "BACKUP__"
SHEETS_MAX_TITLE_LEN = 100

TEMA_MH = {
    "bg": "#f5f7f8",
    "panel": "#ffffff",
    "panel_soft": "#eef3f1",
    "text": "#13221d",
    "muted": "#5f6d67",
    "accent": "#246b47",
    "accent_alt": "#7c8c5a",
    "danger": "#bf3f3f",
    "ok": "#1f7a4b",
    "warning": "#9a6a1f",
    "border": "#d4ddd9",
}

# Tipos monitorados no dominio_dmls_08.py em CONTAS_PASSIVO_PAGAR_CONTROLADAS.
TIPOS_CONTA_PAGAR = {
    "2.1.4.01.003": "ISS A RECOLHER",
    "2.1.4.01.006": "IMPOSTOS LUCRO REAL A RECOLHER",
    "2.1.4.01.015": "SIMPLES NACIONAL A RECOLHER",
    "2.1.5.01.001": "SALARIOS E ORDENADOS A PAGAR",
    "2.1.5.02.001": "INSS A RECOLHER",
    "2.1.5.02.002": "FGTS A RECOLHER",
    "2.1.6.02.001": "HONORARIOS CONTABEIS",
}

COLUNAS_BASE = [
    "empresa",
    "competencia",
    "tipo",
    "descricao",
    "fornecedor_cliente",
    "vencimento",
    "pagamento_recebimento",
    "valor",
    "status",
    "categoria",
    "observacao",
    "documento",
]

COLUNAS_EXTRAS = [
    "tipo_conta_codigo",
    "tipo_conta_nome",
    "anexo_nome",
    "anexo_caminho",
    "codigo_pagamento",
    "criado_em",
    "criado_por",
    "excluido_em",
    "excluido_por",
    "ativo",
]

COLUNAS_DADOS = COLUNAS_BASE + COLUNAS_EXTRAS

COLUNAS_TEXTO = [
    "empresa",
    "competencia",
    "tipo",
    "descricao",
    "fornecedor_cliente",
    "vencimento",
    "pagamento_recebimento",
    "status",
    "categoria",
    "observacao",
    "documento",
    "tipo_conta_codigo",
    "tipo_conta_nome",
    "anexo_nome",
    "anexo_caminho",
    "codigo_pagamento",
    "criado_em",
    "criado_por",
    "excluido_em",
    "excluido_por",
]

COLUNAS_IMPORTACAO = [
    "descricao",
    "fornecedor",
    "vencimento",
    "valor",
    "status",
    "categoria",
    "documento",
    "tipo_conta_codigo",
    "codigo_pagamento",
    "observacao",
]

COLUNAS_IMPORTACAO_OBRIGATORIAS = ["descricao", "fornecedor", "vencimento", "valor"]

DOCUMENTOS_DEMO_LEGADOS = {
    "AP-MHL-001",
    "AP-MHL-002",
    "AP-MHL-003",
    "AP-MHB-001",
    "AP-MHB-002",
}

LIMIAR_QUEDA_SOBRESCRITA = 0.20
VALOR_RELEVANTE_SOBRESCRITA = 1000.0

SUGESTOES_DESCRICAO = [
    "Honorários contábeis",
    "DAS Simples Nacional",
    "INSS a recolher",
    "FGTS a recolher",
    "ISS a recolher",
    "Aluguel",
    "Energia elétrica",
    "Internet",
    "Manutenção de frota",
    "Serviços administrativos",
]

SUGESTOES_FORNECEDOR = [
    "Escritório Contábil",
    "Receita Federal",
    "Prefeitura",
    "Caixa Econômica Federal",
    "Companhia de Energia",
    "Fornecedor Brasil",
]

SUGESTOES_CATEGORIA = [
    "Contabilidade",
    "Tributos",
    "Folha",
    "Administrativo",
    "Operacional",
    "Despesas fixas",
    "Tecnologia",
]

OPCAO_NOVO = "Novo"
OPCOES_OCULTAS = {"outro", "decessars monteiro"}

MODULOS_CONTABEIS = [
    {"id": "contas_a_pagar", "titulo": "Contas a Pagar", "ativo": True},
    {"id": "contas_a_receber", "titulo": "Contas a Receber", "ativo": False},
    {"id": "conciliacao_bancaria", "titulo": "Conciliação Bancária", "ativo": False},
    {"id": "extratos_lancamentos", "titulo": "Extratos e Lançamentos", "ativo": False},
    {"id": "receitas_notas", "titulo": "Receitas e Notas Fiscais", "ativo": False},
    {"id": "impostos_guias", "titulo": "Impostos e Guias", "ativo": False},
    {"id": "folha_inss_fgts", "titulo": "Folha, INSS e FGTS", "ativo": False},
    {"id": "plano_contas", "titulo": "Plano de Contas", "ativo": False},
    {"id": "centros_custo", "titulo": "Centros de Custo", "ativo": False},
    {"id": "balancete", "titulo": "Balancete", "ativo": False},
    {"id": "dre", "titulo": "DRE", "ativo": False},
    {"id": "balanco_patrimonial", "titulo": "Balanço Patrimonial", "ativo": False},
    {"id": "razao_contabil", "titulo": "Razão Contábil", "ativo": False},
    {"id": "relatorios", "titulo": "Relatórios PDF/Excel", "ativo": False},
    {"id": "backup_auditoria", "titulo": "Backup e Auditoria", "ativo": False},
]


def configurar_pagina() -> None:
    st.set_page_config(
        page_title="Portal Contábil do Cliente",
        page_icon="📊",
        layout="wide",
    )

    st.markdown(
        f"""
        <style>
            :root {{
                --mh-bg: {TEMA_MH["bg"]};
                --mh-panel: {TEMA_MH["panel"]};
                --mh-panel-soft: {TEMA_MH["panel_soft"]};
                --mh-text: {TEMA_MH["text"]};
                --mh-muted: {TEMA_MH["muted"]};
                --mh-accent: {TEMA_MH["accent"]};
                --mh-accent-alt: {TEMA_MH["accent_alt"]};
                --mh-border: {TEMA_MH["border"]};
                --mh-danger: {TEMA_MH["danger"]};
                --mh-ok: {TEMA_MH["ok"]};
                --mh-warning: {TEMA_MH["warning"]};
            }}
            [data-testid="stAppViewContainer"] {{
                background:
                    radial-gradient(circle at top left, rgba(36, 107, 71, 0.08), transparent 28rem),
                    linear-gradient(180deg, var(--mh-bg) 0%, #ffffff 72%);
                color: var(--mh-text);
            }}
            [data-testid="stSidebar"] {{
                background: linear-gradient(180deg, #ffffff 0%, var(--mh-panel-soft) 100%);
                border-right: 1px solid var(--mh-border);
            }}
            [data-testid="stSidebar"] img {{
                max-width: 132px;
                display: block;
                margin: 0 auto 0.75rem;
            }}
            .main .block-container {{
                padding-top: 0.7rem;
                padding-bottom: 1.2rem;
            }}
            .portal-header {{
                display: flex;
                align-items: center;
                justify-content: space-between;
                gap: 1.25rem;
                padding: 0.78rem 0.95rem;
                background: var(--mh-panel);
                border: 1px solid var(--mh-border);
                border-radius: 8px;
                box-shadow: 0 8px 22px rgba(19, 34, 29, 0.06);
                margin-bottom: 0.72rem;
            }}
            .portal-brand {{
                display: flex;
                align-items: center;
                gap: 1rem;
                min-width: 0;
            }}
            .portal-logo {{
                width: 124px;
                max-width: 100%;
                height: auto;
                object-fit: contain;
                display: block;
            }}
            .portal-title {{
                margin: 0;
                font-size: clamp(1.12rem, 1.65vw, 1.52rem);
                line-height: 1.15;
                color: var(--mh-text);
                font-weight: 800;
                letter-spacing: 0;
                word-break: normal;
                overflow-wrap: normal;
                hyphens: none;
            }}
            .portal-subtitle {{
                margin: 0.25rem 0 0;
                color: var(--mh-muted);
                font-size: 0.9rem;
            }}
            .portal-meta {{
                display: flex;
                flex-wrap: wrap;
                justify-content: flex-end;
                gap: 0.45rem;
                min-width: 0;
            }}
            .meta-pill {{
                border: 1px solid var(--mh-border);
                border-radius: 8px;
                background: var(--mh-panel-soft);
                color: var(--mh-text);
                padding: 0.42rem 0.62rem;
                font-size: 0.84rem;
                font-weight: 700;
            }}
            .meta-pill span {{
                display: block;
                color: var(--mh-muted);
                font-size: 0.72rem;
                font-weight: 600;
                text-transform: uppercase;
            }}
            .status-ok {{
                border-color: rgba(21, 128, 61, 0.35);
                background: rgba(21, 128, 61, 0.10);
                color: var(--mh-ok);
            }}
            .status-alerta {{
                border-color: rgba(220, 38, 38, 0.35);
                background: rgba(220, 38, 38, 0.10);
                color: var(--mh-danger);
            }}
            .status-pendente {{
                border-color: rgba(183, 121, 31, 0.38);
                background: rgba(183, 121, 31, 0.12);
                color: var(--mh-warning);
            }}
            [data-testid="stMetric"] {{
                background: var(--mh-panel);
                border: 1px solid var(--mh-border);
                border-radius: 8px;
                padding: 14px 16px;
                box-shadow: 0 6px 16px rgba(19, 34, 29, 0.05);
            }}
            [data-testid="stMetricValue"] {{
                color: var(--mh-accent);
            }}
            .metric-grid {{
                display: grid;
                gap: 0.6rem;
                margin: 0.6rem 0 0.78rem;
            }}
            .metric-row {{
                display: grid;
                gap: 0.6rem;
                margin-bottom: 0.6rem;
            }}
            .metric-row.top {{
                grid-template-columns: repeat(4, minmax(0, 1fr));
            }}
            .metric-row.bottom {{
                grid-template-columns: repeat(3, minmax(0, 1fr));
            }}
            .metric-card {{
                background: var(--mh-panel);
                border: 1px solid var(--mh-border);
                border-radius: 8px;
                padding: 0.75rem 0.85rem;
                min-width: 0;
                box-shadow: 0 6px 16px rgba(19, 34, 29, 0.05);
            }}
            .metric-label {{
                color: var(--mh-muted);
                font-size: 0.82rem;
                font-weight: 700;
                margin-bottom: 0.28rem;
                display: flex;
                align-items: center;
                gap: 0.45rem;
            }}
            .metric-value {{
                color: var(--mh-accent);
                font-size: clamp(1.08rem, 2.3vw, 1.55rem);
                line-height: 1.15;
                font-weight: 500;
                white-space: nowrap;
            }}
            .metric-card.metric-danger {{
                border-color: rgba(191, 63, 63, 0.28);
                background: linear-gradient(180deg, rgba(191, 63, 63, 0.08), #ffffff 56%);
            }}
            .metric-card.metric-warning {{
                border-color: rgba(154, 106, 31, 0.28);
                background: linear-gradient(180deg, rgba(154, 106, 31, 0.08), #ffffff 56%);
            }}
            .metric-card.metric-ok {{
                border-color: rgba(31, 122, 75, 0.28);
                background: linear-gradient(180deg, rgba(31, 122, 75, 0.08), #ffffff 56%);
            }}
            .metric-card.metric-info {{
                border-color: rgba(36, 107, 71, 0.22);
                background: linear-gradient(180deg, rgba(36, 107, 71, 0.06), #ffffff 56%);
            }}
            .metric-card.metric-neutral {{
                border-color: rgba(95, 109, 103, 0.18);
            }}
            .metric-delta {{
                display: inline-flex;
                align-items: center;
                width: fit-content;
                margin-top: 0.55rem;
                border-radius: 999px;
                background: rgba(21, 128, 61, 0.10);
                color: var(--mh-ok);
                padding: 0.12rem 0.45rem;
                font-size: 0.85rem;
                font-weight: 800;
            }}
    .login-logo {{
        display: flex;
        justify-content: center;
        margin-bottom: 0.45rem;
    }}
            .login-logo img {{
                width: 128px;
                max-width: 100%;
                height: auto;
                object-fit: contain;
                display: block;
            }}
    .login-panel {{
        border: 1px solid var(--mh-border);
        border-radius: 10px;
        background: var(--mh-panel);
        box-shadow: 0 14px 30px rgba(19, 34, 29, 0.08);
        padding: 0.9rem 1rem;
        text-align: center;
    }}
            .login-panel h2 {{
                margin: 0 0 0.22rem;
                font-size: 1.28rem;
            }}
            .login-panel p {{
                color: var(--mh-muted);
                margin: 0.18rem 0;
                line-height: 1.35;
            }}
            .login-badge {{
                display: inline-flex;
                width: fit-content;
                border-radius: 8px;
                background: rgba(47, 143, 91, 0.12);
                color: var(--mh-accent);
                border: 1px solid rgba(47, 143, 91, 0.24);
                padding: 0.42rem 0.62rem;
                font-size: 0.82rem;
                font-weight: 800;
            }}
            .empresa-card {{
                border: 1px solid var(--mh-border);
                border-radius: 14px;
                background: var(--mh-panel);
                box-shadow: 0 14px 28px rgba(19, 34, 29, 0.08);
                padding: 1rem 1rem 0.9rem;
                min-height: 10rem;
                position: relative;
                overflow: hidden;
            }}
            .empresa-card::before {{
                content: "";
                position: absolute;
                inset: 0 auto auto 0;
                width: 100%;
                height: 6px;
                background: linear-gradient(90deg, var(--mh-accent) 0%, var(--mh-accent-alt) 100%);
            }}
            .empresa-card h3 {{
                margin: 0 0 0.55rem;
                font-size: 1.18rem;
            }}
            .empresa-card p {{
                margin: 0.2rem 0;
                color: var(--mh-muted);
                font-size: 0.9rem;
            }}
            .empresa-card-topo {{
                display: flex;
                align-items: flex-start;
                justify-content: space-between;
                gap: 0.75rem;
                margin-bottom: 0.65rem;
            }}
            .empresa-card-badges {{
                display: flex;
                flex-wrap: wrap;
                gap: 0.45rem;
                margin: 0.7rem 0 0.8rem;
            }}
            .empresa-chip {{
                display: inline-flex;
                align-items: center;
                gap: 0.35rem;
                border-radius: 999px;
                padding: 0.28rem 0.55rem;
                font-size: 0.78rem;
                font-weight: 700;
                border: 1px solid transparent;
            }}
            .empresa-chip.total {{
                background: rgba(47, 143, 91, 0.10);
                color: var(--mh-accent);
                border-color: rgba(47, 143, 91, 0.18);
            }}
            .empresa-chip.abertas {{
                background: rgba(29, 78, 216, 0.08);
                color: #1d4ed8;
                border-color: rgba(29, 78, 216, 0.16);
            }}
            .empresa-chip.vencidas {{
                background: rgba(185, 28, 28, 0.09);
                color: #b91c1c;
                border-color: rgba(185, 28, 28, 0.16);
            }}
            .empresa-chip.info {{
                background: rgba(124, 140, 90, 0.11);
                color: var(--mh-accent-alt);
                border-color: rgba(124, 140, 90, 0.16);
            }}
            .dashboard-shell {{
                max-width: 1180px;
                margin: 0 auto;
            }}
            .section-panel {{
                border: 1px solid var(--mh-border);
                border-radius: 8px;
                background: var(--mh-panel);
                box-shadow: 0 8px 20px rgba(19, 34, 29, 0.05);
                padding: 0.95rem;
                margin: 0.55rem 0 0.85rem;
            }}
            .conta-card {{
                border: 1px solid var(--mh-border);
                border-radius: 10px;
                background: var(--mh-panel);
                box-shadow: 0 8px 18px rgba(19, 34, 29, 0.05);
                padding: 0.85rem 0.9rem;
                margin-bottom: 0.7rem;
            }}
            .conta-card-topo {{
                display: flex;
                align-items: flex-start;
                justify-content: space-between;
                gap: 0.75rem;
                margin-bottom: 0.7rem;
            }}
            .conta-card-documento {{
                color: var(--mh-text);
                font-size: 0.98rem;
                font-weight: 800;
                word-break: normal;
                overflow-wrap: anywhere;
            }}
            .conta-card-status {{
                display: inline-flex;
                align-items: center;
                justify-content: center;
                border-radius: 999px;
                padding: 0.28rem 0.55rem;
                font-size: 0.75rem;
                font-weight: 800;
                white-space: nowrap;
                border: 1px solid transparent;
            }}
            .conta-status-vencida {{
                background: rgba(220, 38, 38, 0.10);
                color: var(--mh-danger);
                border-color: rgba(220, 38, 38, 0.22);
            }}
            .conta-status-atencao {{
                background: rgba(183, 121, 31, 0.12);
                color: var(--mh-warning);
                border-color: rgba(183, 121, 31, 0.24);
            }}
            .conta-status-ok {{
                background: rgba(21, 128, 61, 0.10);
                color: var(--mh-ok);
                border-color: rgba(21, 128, 61, 0.24);
            }}
            .conta-status-neutra {{
                background: rgba(95, 109, 103, 0.10);
                color: var(--mh-text);
                border-color: rgba(95, 109, 103, 0.18);
            }}
            .conta-card-grid {{
                display: grid;
                grid-template-columns: repeat(2, minmax(0, 1fr));
                gap: 0.6rem 0.85rem;
            }}
            .conta-card-item {{
                min-width: 0;
            }}
            .conta-card-label {{
                display: block;
                margin-bottom: 0.22rem;
                color: var(--mh-muted);
                font-size: 0.74rem;
                font-weight: 700;
                text-transform: uppercase;
                letter-spacing: 0.02em;
            }}
            .conta-card-value {{
                color: var(--mh-text);
                font-size: 0.92rem;
                line-height: 1.25;
                word-break: normal;
                overflow-wrap: anywhere;
            }}
            .conta-card-valor {{
                color: var(--mh-accent);
                font-weight: 800;
            }}
            [class*="st-key-contas_card_"] button {{
                min-height: 2rem !important;
                min-width: 2rem !important;
                padding: 0.2rem 0.3rem !important;
                font-size: 0.95rem !important;
                line-height: 1 !important;
            }}
            .small-muted {{
                color: var(--mh-muted);
                font-size: 0.9rem;
            }}
            .menu-contabil-titulo {{
                color: var(--mh-text);
                font-weight: 800;
                font-size: 0.98rem;
                margin: 0.4rem 0 0.25rem;
            }}
            .menu-contabil-ajuda {{
                color: var(--mh-muted);
                font-size: 0.78rem;
                margin: 0 0 0.45rem;
            }}
            div.stButton > button,
            div.stDownloadButton > button,
            div.stFormSubmitButton > button {{
                border-radius: 8px;
                border: 1px solid var(--mh-accent);
                background: var(--mh-accent);
                color: #ffffff;
                font-weight: 700;
                min-height: 2.2rem;
                padding: 0.35rem 0.8rem;
                white-space: nowrap;
                overflow: hidden;
                text-overflow: ellipsis;
            }}
            [class*="st-key-ordenar_contas_"] button,
            [class*="st-key-contas_acao_"] button,
            [class*="st-key-baixar_excel_contas"] button {{
                border-color: var(--mh-border) !important;
                background: var(--mh-panel-soft) !important;
                color: var(--mh-text) !important;
                box-shadow: none !important;
                font-weight: 700;
                min-height: 1.9rem !important;
                padding: 0.18rem 0.55rem !important;
                white-space: nowrap !important;
                overflow: hidden !important;
                text-overflow: ellipsis !important;
            }}
            [class*="st-key-ordenar_contas_"] button:hover,
            [class*="st-key-contas_acao_"] button:hover,
            [class*="st-key-baixar_excel_contas"] button:hover {{
                border-color: var(--mh-accent-alt) !important;
                background: rgba(36, 107, 71, 0.08) !important;
                color: var(--mh-accent) !important;
            }}
            div.stButton > button:hover,
            div.stDownloadButton > button:hover,
            div.stFormSubmitButton > button:hover {{
                border-color: var(--mh-accent-alt);
                background: var(--mh-accent-alt);
                color: #ffffff;
            }}
            [data-testid="stDataFrame"] {{
                border: 1px solid var(--mh-border);
                border-radius: 8px;
                overflow: hidden;
            }}
            h1, h2, h3 {{
                color: var(--mh-text);
                letter-spacing: 0;
                word-break: normal;
                overflow-wrap: normal;
                hyphens: none;
            }}
            hr {{
                border-color: var(--mh-border);
            }}
            @media (max-width: 760px) {{
                [data-testid="stAppViewContainer"] {{
                    background: linear-gradient(180deg, var(--mh-bg) 0%, #ffffff 78%);
                }}
                .main .block-container {{
                    padding-left: 0.85rem;
                    padding-right: 0.85rem;
                    padding-top: 0.75rem;
                    max-width: 100%;
                }}
                [data-testid="stSidebar"] {{
                    border-right: 0;
                    border-bottom: 1px solid var(--mh-border);
                }}
                [data-testid="stSidebar"] img {{
                    max-width: 104px;
                    margin-bottom: 0.45rem;
                }}
                .portal-header {{
                    align-items: flex-start;
                    flex-direction: column;
                    gap: 0.65rem;
                    padding: 0.78rem 0.85rem;
                }}
                .portal-brand {{
                    align-items: flex-start;
                    flex-direction: column;
                    gap: 0.45rem;
                }}
                .portal-logo {{
                    width: 100px;
                }}
                .portal-meta {{
                    justify-content: flex-start;
                    min-width: 0;
                    width: 100%;
                }}
                .meta-pill {{
                    flex: 1 1 100%;
                    padding: 0.42rem 0.6rem;
                }}
                .portal-title {{
                    font-size: 1.12rem;
                    line-height: 1.2;
                }}
                .portal-subtitle {{
                    font-size: 0.86rem;
                }}
                .login-logo img {{
                    width: 96px;
                }}
                .login-panel {{
                    padding: 0.72rem 0.8rem;
                }}
                .login-panel h2 {{
                    font-size: 1.02rem;
                }}
                .empresa-card {{
                    min-height: 0;
                    padding: 0.9rem 0.9rem 0.8rem;
                    margin-top: 0.2rem;
                }}
                .empresa-card h3 {{
                    font-size: 1rem;
                }}
                .empresa-card p {{
                    font-size: 0.86rem;
                }}
                [data-testid="stMetric"] {{
                    padding: 0.75rem 0.85rem;
                }}
                [data-testid="stMetricLabel"] {{
                    font-size: 0.78rem;
                }}
                [data-testid="stMetricValue"] {{
                    font-size: 1.15rem;
                }}
                .metric-grid {{
                    gap: 0.55rem;
                }}
                .metric-row.top,
                .metric-row.bottom {{
                    grid-template-columns: 1fr;
                }}
                .conta-card-grid {{
                    grid-template-columns: 1fr;
                }}
                .metric-card {{
                    padding: 0.72rem 0.82rem;
                }}
                .metric-label {{
                    font-size: 0.8rem;
                }}
                .metric-value {{
                    font-size: clamp(1rem, 5vw, 1.35rem);
                }}
                div.stButton > button,
                div.stDownloadButton > button,
                div.stFormSubmitButton > button {{
                    min-height: 2.45rem;
                    font-size: 0.92rem;
                }}
                [class*="st-key-ordenar_contas_"] button,
                [class*="st-key-contas_acao_"] button,
                [class*="st-key-baixar_excel_contas"] button {{
                    min-height: 1.85rem !important;
                    padding: 0.2rem 0.5rem !important;
                }}
                [class*="st-key-contas_card_"] button {{
                    min-height: 1.9rem !important;
                    min-width: 1.9rem !important;
                    padding: 0.18rem 0.28rem !important;
                }}
                [data-testid="stDataFrame"] {{
                    overflow-x: auto;
                }}
                h1 {{
                    font-size: 1.7rem;
                    line-height: 1.2;
                }}
                h2 {{
                    font-size: 1.35rem;
                }}
                h3 {{
                    font-size: 1.12rem;
                }}
            }}
            @media (max-width: 1200px) {{
                .portal-header {{
                    align-items: flex-start;
                    flex-direction: column;
                    gap: 0.65rem;
                }}
                .portal-brand {{
                    width: 100%;
                    align-items: flex-start;
                    flex-wrap: wrap;
                    gap: 0.6rem;
                }}
                .portal-meta {{
                    width: 100%;
                    justify-content: flex-start;
                }}
                .metric-row.top,
                .metric-row.bottom {{
                    grid-template-columns: repeat(2, minmax(0, 1fr));
                }}
            }}
            @media (max-width: 900px) {{
                .metric-row.top,
                .metric-row.bottom {{
                    grid-template-columns: 1fr;
                }}
            }}
        </style>
        """,
        unsafe_allow_html=True,
    )


def imagem_base64(caminho: Path) -> str:
    if not caminho.exists():
        return ""
    return base64.b64encode(caminho.read_bytes()).decode("ascii")


def mostrar_cabecalho(empresa: str, status_geral: str, status_tipo: str) -> None:
    logo_b64 = imagem_base64(LOGO_FULL_PATH)
    logo_html = (
        f'<img class="portal-logo" src="data:image/png;base64,{logo_b64}" alt="MH LOG">'
        if logo_b64
        else '<strong style="color:var(--mh-accent);font-size:1.5rem;">MH LOG</strong>'
    )
    status_classe = {
        "ok": "status-ok",
        "alerta": "status-alerta",
        "pendente": "status-pendente",
    }.get(status_tipo, "status-pendente")

    st.markdown(
        f"""
        <section class="portal-header">
            <div class="portal-brand">
                <div>{logo_html}</div>
                <div>
                    <h1 class="portal-title">Portal Contábil do Cliente</h1>
                    <p class="portal-subtitle">
                        Contas a pagar liberadas para acompanhamento e registro manual.
                    </p>
                </div>
            </div>
            <div class="portal-meta">
                <div class="meta-pill"><span>Cliente</span>{escape(empresa)}</div>
                <div class="meta-pill {status_classe}"><span>Status</span>{escape(status_geral)}</div>
            </div>
        </section>
        """,
        unsafe_allow_html=True,
    )


def renderizar_metricas(cards: list[dict[str, object]]) -> None:
    itens = []
    for card in cards:
        delta = card.get("delta")
        delta_html = f'<div class="metric-delta">↑ {escape(str(delta))}</div>' if delta is not None else ""
        itens.append(
            '<div class="metric-card">'
            f'<div class="metric-label">{escape(str(card["label"]))}</div>'
            f'<div class="metric-value">{escape(str(card["value"]))}</div>'
            f"{delta_html}"
            "</div>"
        )

    st.markdown(f'<div class="metric-grid">{"".join(itens)}</div>', unsafe_allow_html=True)


def renderizar_metricas_em_duas_linhas(cards: list[dict[str, object]]) -> None:
    def render_card(card: dict[str, object]) -> str:
        delta = card.get("delta")
        emoji = str(card.get("emoji", ""))
        tone = str(card.get("tone", "neutral")).strip().lower()
        delta_html = f'<div class="metric-delta">↑ {escape(str(delta))}</div>' if delta is not None else ""
        return (
            f'<div class="metric-card metric-{escape(tone)}">'
            f'<div class="metric-label"><span>{escape(emoji)}</span><span>{escape(str(card["label"]))}</span></div>'
            f'<div class="metric-value">{escape(str(card["value"]))}</div>'
            f"{delta_html}"
            "</div>"
        )

    topo = cards[:4]
    base = cards[4:]
    topo_html = "".join(render_card(card) for card in topo)
    base_html = "".join(render_card(card) for card in base)
    st.markdown(
        f"""
        <div class="metric-grid">
            <div class="metric-row top">{topo_html}</div>
            <div class="metric-row bottom">{base_html}</div>
        </div>
        """,
        unsafe_allow_html=True,
    )


def criar_config_demo() -> dict:
    config = {
        "usuarios": [
            {"usuario": "DMLIMA", "senha": "123456", "empresas": ["MHLOG", "MH BRASIL"]},
            {"usuario": "VICTOR", "senha": "123456", "empresas": ["MHLOG", "MH BRASIL"]},
            {"usuario": "ALEX", "senha": "123456", "empresas": ["MHLOG", "MH BRASIL"]},
        ],
        "clientes": [
            {"empresa": "MHLOG"},
            {"empresa": "MH BRASIL"},
        ]
    }
    CONFIG_PATH.write_text(json.dumps(config, indent=2, ensure_ascii=False), encoding="utf-8")
    return config


def normalizar_config(config: dict) -> dict:
    clientes = config.get("clientes", [])
    empresas = [cliente.get("empresa", "") for cliente in clientes if cliente.get("empresa")]
    usuarios = config.get("usuarios", [])

    if not usuarios:
        usuarios = [
            {
                "usuario": item.get("empresa", ""),
                "senha": item.get("senha", ""),
                "empresas": empresas,
            }
            for item in clientes
            if item.get("senha")
        ]

    for cliente in clientes:
        cliente.pop("senha", None)

    return {"usuarios": usuarios, "clientes": clientes}


def carregar_config() -> dict:
    if not CONFIG_PATH.exists():
        return criar_config_demo()

    with CONFIG_PATH.open("r", encoding="utf-8") as arquivo:
        return normalizar_config(json.load(arquivo))


def dados_demo() -> pd.DataFrame:
    return pd.DataFrame(columns=COLUNAS_DADOS)


def criar_dados_demo() -> pd.DataFrame:
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    df = dados_demo()
    df.to_csv(CSV_PATH, index=False, encoding="utf-8-sig")
    return df


def normalizar_ativo(valor: object) -> bool:
    if isinstance(valor, bool):
        return valor
    texto = str(valor).strip().lower()
    return texto not in {"false", "0", "nao", "não", "n", "excluido", "excluído"}


def remover_linhas_em_branco(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df.copy()

    colunas_verificacao = [coluna for coluna in COLUNAS_DADOS if coluna != "ativo" and coluna in df.columns]
    if not colunas_verificacao:
        return df.copy()

    vazia = pd.Series(True, index=df.index)
    for coluna in colunas_verificacao:
        serie = df[coluna]
        vazia &= serie.isna() | serie.astype(str).str.strip().eq("")

    return df.loc[~vazia].copy()


def normalizar_valor_operacional(valor: object) -> float:
    if pd.isna(valor):
        return 0.0
    if isinstance(valor, (int, float)):
        return float(valor)

    texto = str(valor).strip()
    if not texto:
        return 0.0
    texto = texto.replace("R$", "").replace(" ", "")
    if "," in texto:
        texto = texto.replace(".", "").replace(",", ".")
    numero = pd.to_numeric(texto, errors="coerce")
    return float(numero) if not pd.isna(numero) else 0.0


def normalizar_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    df = df.loc[:, ~df.columns.duplicated()].copy()
    df = remover_linhas_em_branco(df)
    for coluna in COLUNAS_DADOS:
        if coluna not in df.columns:
            df[coluna] = True if coluna == "ativo" else ""

    df = df[COLUNAS_DADOS].copy()
    df["valor"] = df["valor"].apply(normalizar_valor_operacional)
    for coluna in COLUNAS_TEXTO:
        df[coluna] = df[coluna].fillna("").astype(str)
    df["ativo"] = df["ativo"].apply(normalizar_ativo)
    df["vencimento_dt"] = pd.to_datetime(df["vencimento"], errors="coerce")
    df["pagamento_recebimento_dt"] = pd.to_datetime(df["pagamento_recebimento"], errors="coerce")
    return df


def slug_empresa(empresa: str) -> str:
    slug = re.sub(r"[^A-Za-z0-9]+", "_", empresa.strip().upper()).strip("_")
    return slug or "EMPRESA"


def pasta_empresa(empresa: str) -> Path:
    return EMPRESAS_DATA_DIR / slug_empresa(empresa)


def caminho_dados_empresa(empresa: str) -> Path:
    return pasta_empresa(empresa) / "dados_portal.csv"


def usar_google_sheets() -> bool:
    try:
        configuracao = st.secrets.get("google_sheets", {})
        credenciais = st.secrets.get("google_service_account", {})
    except Exception:
        return False
    return bool(str(configuracao.get("spreadsheet_id", "")).strip() and credenciais)


def obter_config_google_sheets() -> dict[str, object]:
    try:
        configuracao = dict(st.secrets.get("google_sheets", {}))
        credenciais = dict(st.secrets.get("google_service_account", {}))
    except Exception as erro:
        raise RuntimeError("Configuração do Google Sheets indisponível em st.secrets.") from erro

    spreadsheet_id = str(configuracao.get("spreadsheet_id", "")).strip()
    if not spreadsheet_id:
        raise RuntimeError("Defina google_sheets.spreadsheet_id em st.secrets.")
    if not credenciais:
        raise RuntimeError("Defina google_service_account em st.secrets.")

    return {
        "spreadsheet_id": spreadsheet_id,
        "service_account": credenciais,
        "sheet_prefix": str(configuracao.get("sheet_prefix", "")).strip(),
    }


def nome_aba_empresa(empresa: str) -> str:
    return slug_empresa(empresa)[:SHEETS_MAX_TITLE_LEN]


def nome_aba_backup(empresa: str, timestamp: str | None = None) -> str:
    momento = timestamp or datetime.now().strftime("%Y%m%d_%H%M%S")
    nome = f"{BACKUP_SHEET_PREFIX}{slug_empresa(empresa)}__{momento}"
    return nome[:SHEETS_MAX_TITLE_LEN]


def nome_aba_backup_prefixo(empresa: str) -> str:
    return f"{BACKUP_SHEET_PREFIX}{slug_empresa(empresa)}__"


@st.cache_resource(show_spinner=False)
def cliente_google_sheets() -> object:
    if not usar_google_sheets():
        raise RuntimeError("Google Sheets nao configurado.")

    try:
        import gspread
        from google.oauth2.service_account import Credentials
    except ImportError as erro:
        raise RuntimeError(
            "Instale gspread e google-auth para usar Google Sheets como persistencia."
        ) from erro

    configuracao = obter_config_google_sheets()
    scopes = [
        "https://www.googleapis.com/auth/spreadsheets",
        "https://www.googleapis.com/auth/drive",
    ]
    credenciais = Credentials.from_service_account_info(configuracao["service_account"], scopes=scopes)
    return gspread.authorize(credenciais)


def planilha_google() -> object:
    cliente = cliente_google_sheets()
    configuracao = obter_config_google_sheets()
    return cliente.open_by_key(str(configuracao["spreadsheet_id"]))


def google_sheets_operacional_disponivel(empresa: str | None = None) -> bool:
    if not usar_google_sheets():
        return False
    try:
        if empresa is None:
            planilha_google()
        else:
            obter_worksheet_empresa(empresa, criar=False)
        return True
    except Exception:
        return False


def worksheet_empresa_existe(empresa: str) -> bool:
    if not usar_google_sheets():
        return caminho_dados_empresa(empresa).exists()
    try:
        planilha = planilha_google()
        nome = nome_aba_empresa(empresa)
        return any(aba.title == nome for aba in planilha.worksheets())
    except Exception:
        return caminho_dados_empresa(empresa).exists()


def obter_worksheet_empresa(empresa: str, criar: bool = False):
    if not usar_google_sheets():
        return None

    planilha = planilha_google()
    nome = nome_aba_empresa(empresa)
    try:
        return planilha.worksheet(nome)
    except Exception:
        if not criar:
            return None
        aba = planilha.add_worksheet(title=nome, rows=max(200, len(COLUNAS_DADOS) + 10), cols=len(COLUNAS_DADOS))
        aba.update([COLUNAS_DADOS], value_input_option="RAW")
        return aba


def converter_dataframe_para_linhas(df: pd.DataFrame) -> list[list[object]]:
    dados = normalizar_dataframe(df)
    dados = dados[[coluna for coluna in COLUNAS_DADOS if coluna in dados.columns]].copy()
    for coluna in COLUNAS_DADOS:
        if coluna not in dados.columns:
            dados[coluna] = ""
    dados = dados[COLUNAS_DADOS].copy()

    linhas: list[list[object]] = [COLUNAS_DADOS]
    for _, linha in dados.iterrows():
        valores = []
        for coluna in COLUNAS_DADOS:
            valor = linha.get(coluna, "")
            if pd.isna(valor):
                valor = ""
            elif isinstance(valor, bool):
                valor = bool(valor)
            elif isinstance(valor, (pd.Timestamp, datetime)):
                valor = valor.strftime("%Y-%m-%d %H:%M:%S")
            valores.append(valor)
        linhas.append(valores)
    return linhas


def dataframe_from_linhas(linhas: list[list[object]]) -> pd.DataFrame:
    if not linhas:
        return normalizar_dataframe(pd.DataFrame(columns=COLUNAS_DADOS))

    cabecalho = [str(item).strip() for item in linhas[0]]
    registros = []
    for linha in linhas[1:]:
        if not any(str(valor).strip() for valor in linha):
            continue
        registro = {}
        for indice, coluna in enumerate(COLUNAS_DADOS):
            valor = linha[indice] if indice < len(linha) else ""
            registro[coluna] = valor
        registros.append(registro)
    return normalizar_dataframe(pd.DataFrame(registros, columns=COLUNAS_DADOS))


def ler_df_empresa_google(empresa: str) -> pd.DataFrame:
    aba = obter_worksheet_empresa(empresa, criar=False)
    if aba is None:
        return normalizar_dataframe(pd.DataFrame(columns=COLUNAS_DADOS))
    valores = aba.get_all_values()
    if not valores:
        return normalizar_dataframe(pd.DataFrame(columns=COLUNAS_DADOS))
    return dataframe_from_linhas(valores)


def migrar_csv_local_para_google_sheets(empresa: str) -> pd.DataFrame:
    caminho = caminho_dados_empresa(empresa)
    if not caminho.exists():
        return normalizar_dataframe(pd.DataFrame(columns=COLUNAS_DADOS))

    df_local = normalizar_dataframe(pd.read_csv(caminho))
    if df_local.empty:
        return df_local

    aba = obter_worksheet_empresa(empresa, criar=True)
    aba.clear()
    aba.update(converter_dataframe_para_linhas(df_local), value_input_option="RAW")
    return df_local


def listar_backups_empresa_google(empresa: str) -> list[str]:
    try:
        planilha = planilha_google()
        prefixo = nome_aba_backup_prefixo(empresa)
        backups = [aba.title for aba in planilha.worksheets() if aba.title.startswith(prefixo)]
        return sorted(backups, reverse=True)
    except Exception:
        return []


def criar_backup_empresa_google(empresa: str) -> str | None:
    try:
        aba_origem = obter_worksheet_empresa(empresa, criar=False)
        if aba_origem is None:
            return None

        valores = aba_origem.get_all_values()
        if not valores:
            return None

        planilha = planilha_google()
        nome_backup = nome_aba_backup(empresa)
        contador = 1
        while any(aba.title == nome_backup for aba in planilha.worksheets()):
            nome_backup = f"{BACKUP_SHEET_PREFIX}{slug_empresa(empresa)}__{datetime.now().strftime('%Y%m%d_%H%M%S')}_{contador}"
            nome_backup = nome_backup[:SHEETS_MAX_TITLE_LEN]
            contador += 1

        rows = max(2, len(valores))
        cols = max(2, len(valores[0]) if valores else len(COLUNAS_DADOS))
        aba_backup = planilha.add_worksheet(title=nome_backup, rows=rows, cols=cols)
        aba_backup.update(valores, value_input_option="RAW")
        return nome_backup
    except Exception:
        return None


def restaurar_backup_empresa_google(empresa: str, backup: str | None = None) -> str:
    try:
        planilha = planilha_google()
        aba_backup_nome = backup or (listar_backups_empresa_google(empresa)[0] if listar_backups_empresa_google(empresa) else "")
        if not aba_backup_nome:
            raise ValueError("Nenhum backup encontrado para restauracao.")

        aba_backup = planilha.worksheet(aba_backup_nome)
        valores = aba_backup.get_all_values()
        if not valores:
            raise ValueError("O backup selecionado esta vazio.")

        aba_empresa = obter_worksheet_empresa(empresa, criar=True)
        aba_empresa.clear()
        aba_empresa.update(valores, value_input_option="RAW")
        return aba_empresa.title
    except Exception as erro:
        raise ValueError(f"Nao foi possivel restaurar o backup na planilha: {erro}") from erro


def identificar_base_operacional(empresa: str) -> str:
    if google_sheets_operacional_disponivel(empresa):
        return f"Google Sheets / {nome_aba_empresa(empresa)}"
    return str(caminho_dados_empresa(empresa))


def pasta_backups_empresa(empresa: str) -> Path:
    return BACKUPS_DIR / slug_empresa(empresa)


def listar_backups_empresa(empresa: str) -> list[Path | str]:
    if google_sheets_operacional_disponivel(empresa):
        return listar_backups_empresa_google(empresa)

    pasta = pasta_backups_empresa(empresa)
    if not pasta.exists():
        return []
    backups = [caminho for caminho in pasta.glob("dados_portal_*.csv") if caminho.is_file()]
    return sorted(backups, key=lambda caminho: caminho.stat().st_mtime, reverse=True)


def criar_backup_empresa(empresa: str, origem: Path | None = None) -> Path | str | None:
    if google_sheets_operacional_disponivel(empresa):
        nome_backup = criar_backup_empresa_google(empresa)
        return nome_backup

    origem = origem or caminho_dados_empresa(empresa)
    if not origem.exists():
        return None

    pasta = pasta_backups_empresa(empresa)
    pasta.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    destino = pasta / f"dados_portal_{timestamp}.csv"
    contador = 1
    while destino.exists():
        destino = pasta / f"dados_portal_{timestamp}_{contador}.csv"
        contador += 1
    shutil.copy2(origem, destino)
    return destino


def obter_resumo_base_empresa(empresa: str, df_base: pd.DataFrame | None = None) -> dict:
    caminho = identificar_base_operacional(empresa)
    backups = listar_backups_empresa(empresa)
    resumo = {
        "empresa": empresa,
        "caminho": caminho,
        "existe": worksheet_empresa_existe(empresa),
        "status": "ok",
        "erro": "",
        "qtd_registros": 0,
        "qtd_contas_ativas": 0,
        "valor_contas_ativas": 0.0,
        "qtd_backups": len(backups),
        "ultimo_backup": backups[0] if backups else None,
        "backups": backups,
    }

    if df_base is None:
        if not worksheet_empresa_existe(empresa):
            resumo["status"] = "ausente"
            return resumo
        try:
            df_base = carregar_dados_empresa(empresa)
        except Exception as erro:
            resumo["status"] = "erro"
            resumo["erro"] = str(erro)
            return resumo

    if not worksheet_empresa_existe(empresa):
        resumo["status"] = "ausente"

    dados = normalizar_dataframe(df_base)
    resumo["qtd_registros"] = int(len(dados))
    contas_ativas = dados.loc[(dados["tipo"].astype(str) == "conta_a_pagar") & dados["ativo"]].copy()
    resumo["qtd_contas_ativas"] = int(len(contas_ativas))
    resumo["valor_contas_ativas"] = float(pd.to_numeric(contas_ativas["valor"], errors="coerce").fillna(0).sum()) if not contas_ativas.empty else 0.0
    caminho_local = caminho_dados_empresa(empresa)
    if caminho_local.exists() and not resumo["qtd_registros"]:
        resumo["status"] = "vazia"
    return resumo


def base_operacional_suspeita(resumo_atual: dict, df_novo: pd.DataFrame) -> tuple[bool, str]:
    dados_novos = normalizar_dataframe(df_novo)
    novas_ativas = dados_novos.loc[(dados_novos["tipo"].astype(str) == "conta_a_pagar") & dados_novos["ativo"]].copy()
    qtd_novas = int(len(novas_ativas))
    valor_novo = float(pd.to_numeric(novas_ativas["valor"], errors="coerce").fillna(0).sum()) if qtd_novas else 0.0

    qtd_atual = int(resumo_atual.get("qtd_contas_ativas", 0) or 0)
    valor_atual = float(resumo_atual.get("valor_contas_ativas", 0) or 0)

    if qtd_atual <= 0 and valor_atual <= 0:
        return False, ""
    if qtd_novas == 0:
        return True, "a nova gravação ficou sem contas ativas"
    if valor_atual >= VALOR_RELEVANTE_SOBRESCRITA and valor_novo <= 0:
        return True, "o valor total ativo caiu para zero"
    if qtd_atual >= 5 and qtd_novas / max(qtd_atual, 1) < LIMIAR_QUEDA_SOBRESCRITA:
        return True, "a quantidade de registros ativos caiu bruscamente"
    if valor_atual >= VALOR_RELEVANTE_SOBRESCRITA and valor_novo / max(valor_atual, 1) < LIMIAR_QUEDA_SOBRESCRITA:
        return True, "o valor total ativo caiu bruscamente"
    return False, ""


def restaurar_backup_empresa(empresa: str, backup: Path | None = None) -> Path | str:
    if google_sheets_operacional_disponivel(empresa):
        backup_nome = str(backup) if backup is not None else None
        try:
            return restaurar_backup_empresa_google(empresa, backup_nome)
        except Exception:
            pass

    backup = backup or obter_resumo_base_empresa(empresa).get("ultimo_backup")
    if not backup:
        raise ValueError("Nenhum backup encontrado para restauracao.")

    backup = Path(backup)
    if not backup.exists():
        raise ValueError("O backup selecionado nao existe mais.")

    caminho = caminho_dados_empresa(empresa)
    caminho.parent.mkdir(parents=True, exist_ok=True)
    if caminho.exists():
        criar_backup_empresa(empresa, caminho)
    shutil.copy2(backup, caminho)
    return caminho


def exibir_alerta_base_operacional(resumo: dict) -> None:
    empresa = str(resumo.get("empresa", ""))
    caminho = resumo.get("caminho")
    quantidade_backups = int(resumo.get("qtd_backups", 0) or 0)
    ultimo_backup = resumo.get("ultimo_backup")
    status = str(resumo.get("status", "ok"))
    fonte = "Planilha" if usar_google_sheets() else "CSV"

    if status == "ok" and int(resumo.get("qtd_registros", 0) or 0) > 0:
        return

    if status == "ausente":
        st.error(
            "Base operacional nao encontrada. A empresa pode ter dados reais em backup, mas o CSV atual nao esta disponivel. Restaure o backup mais recente ou reimporte a base."
        )
    elif status == "vazia":
        st.warning(
            "A base operacional desta empresa esta vazia. Se isso nao era esperado, restaure um backup ou reimporte a base antes de prosseguir."
        )
    elif status == "erro":
        st.error(f"Falha ao ler a base operacional desta empresa: {resumo.get('erro', '')}")
    else:
        return

    st.markdown(
        f"""
        **Empresa:** {escape(empresa)}  
        **{fonte} esperada:** `{caminho}`  
        **Backups encontrados:** {quantidade_backups}
        """,
        unsafe_allow_html=True,
    )

    if ultimo_backup:
        st.caption(f"Backup mais recente: {ultimo_backup}")

    if quantidade_backups:
        if st.button(
            "Restaurar backup mais recente",
            key=f"restaurar_backup_{slug_empresa(empresa)}",
            use_container_width=True,
        ):
            caminho_restaurado = restaurar_backup_empresa(empresa, ultimo_backup)
            st.success(f"Backup restaurado em {caminho_restaurado}.")
            st.rerun()
    else:
        st.info("Nenhum backup localizado para esta empresa. Verifique a pasta data/backups/<EMPRESA>/ antes de gravar uma nova base.")


def renderizar_salvamento_pendente() -> None:
    pendente = st.session_state.get("salvamento_pendente")
    if not pendente:
        return

    empresa = str(pendente.get("empresa", ""))
    motivo = str(pendente.get("motivo", "alteracao suspeita"))
    qtd_atual = int(pendente.get("qtd_atual", 0) or 0)
    valor_atual = float(pendente.get("valor_atual", 0) or 0)
    qtd_nova = int(pendente.get("qtd_nova", 0) or 0)
    valor_novo = float(pendente.get("valor_novo", 0) or 0)

    st.warning(
        f"Gravacao bloqueada para proteger a base de {empresa}: {motivo}. "
        "Confirme explicitamente se essa alteracao for esperada."
    )
    st.write(
        f"Base atual: {qtd_atual} registro(s) ativo(s), {formatar_moeda_br(valor_atual)} | "
        f"Nova base: {qtd_nova} registro(s) ativo(s), {formatar_moeda_br(valor_novo)}"
    )

    c1, c2 = st.columns(2)
    if c1.button("Confirmar gravacao", key=f"confirmar_salvamento_{slug_empresa(empresa)}", use_container_width=True):
        df_pendente = pendente.get("df")
        try:
            salvar_dados_empresa(df_pendente, empresa, confirmar_sobrescrita_perigosa=True)
        except Exception as erro:
            st.error(str(erro))
            return
        st.session_state.pop("salvamento_pendente", None)
        st.success("Gravacao confirmada e salva com backup automatico.")
        st.rerun()
    if c2.button("Cancelar gravacao", key=f"cancelar_salvamento_{slug_empresa(empresa)}", use_container_width=True):
        st.session_state.pop("salvamento_pendente", None)
        st.info("Gravacao pendente cancelada. Nenhum arquivo foi sobrescrito.")
        st.rerun()


def carregar_base_demo() -> pd.DataFrame:
    if CSV_PATH.exists():
        return pd.read_csv(CSV_PATH)
    if XLSX_PATH.exists():
        return pd.read_excel(XLSX_PATH)
    return criar_dados_demo()


def carregar_dados_empresa(empresa: str) -> pd.DataFrame:
    """Carrega a base operacional isolada de uma empresa."""
    caminho = caminho_dados_empresa(empresa)
    if google_sheets_operacional_disponivel(empresa):
        try:
            df = ler_df_empresa_google(empresa)
            if df.empty:
                df = migrar_csv_local_para_google_sheets(empresa)
        except Exception:
            df = pd.read_csv(caminho) if caminho.exists() else pd.DataFrame(columns=COLUNAS_DADOS)
    elif caminho.exists():
        df = pd.read_csv(caminho)
    else:
        df = pd.DataFrame(columns=COLUNAS_DADOS)
    df_normalizado = normalizar_dataframe(df)
    df_limpo = remover_lancamentos_demo_legados(df_normalizado)
    return df_limpo


def remover_lancamentos_demo_legados(df: pd.DataFrame) -> pd.DataFrame:
    documentos = df["documento"].astype(str).str.strip()
    criado_em = df["criado_em"].fillna("").astype(str).str.strip()
    mascara_demo = documentos.isin(DOCUMENTOS_DEMO_LEGADOS) & criado_em.eq("")
    if not mascara_demo.any():
        return df
    return df.loc[~mascara_demo].copy()


def carregar_dados_empresas(config: dict) -> pd.DataFrame:
    frames = [carregar_dados_empresa(cliente["empresa"]) for cliente in config["clientes"]]
    if not frames:
        return normalizar_dataframe(pd.DataFrame(columns=COLUNAS_DADOS))
    return normalizar_dataframe(pd.concat(frames, ignore_index=True))


def salvar_dados_empresa(
    df: pd.DataFrame,
    empresa: str,
    confirmar_sobrescrita_perigosa: bool = False,
) -> Path | None:
    """Salva no arquivo operacional da empresa selecionada com backup e protecao."""
    caminho = caminho_dados_empresa(empresa)
    caminho.parent.mkdir(parents=True, exist_ok=True)
    dados = normalizar_dataframe(df)
    dados = dados[[coluna for coluna in COLUNAS_DADOS if coluna in dados.columns]].copy()
    resumo_atual = obter_resumo_base_empresa(empresa)
    suspeita, motivo = base_operacional_suspeita(resumo_atual, dados)
    if suspeita and not confirmar_sobrescrita_perigosa:
        st.session_state["salvamento_pendente"] = {
            "empresa": empresa,
            "df": dados.copy(),
            "motivo": motivo,
            "qtd_atual": resumo_atual.get("qtd_contas_ativas", 0),
            "valor_atual": resumo_atual.get("valor_contas_ativas", 0.0),
            "qtd_nova": int(len(dados.loc[(dados["tipo"].astype(str) == "conta_a_pagar") & dados["ativo"]])),
            "valor_novo": float(
                pd.to_numeric(
                    dados.loc[(dados["tipo"].astype(str) == "conta_a_pagar") & dados["ativo"], "valor"],
                    errors="coerce",
                ).fillna(0).sum()
            ),
        }
        st.warning(
            f"Gravacao bloqueada para proteger a base de {empresa}. "
            "Confirme explicitamente no bloco de salvamento pendente para prosseguir."
        )
        return None

    if google_sheets_operacional_disponivel(empresa):
        try:
            criar_backup_empresa(empresa)
            aba = obter_worksheet_empresa(empresa, criar=True)
            valores = converter_dataframe_para_linhas(dados)
            aba.clear()
            aba.update(valores, value_input_option="RAW")
            caminho_saida: Path | str = identificar_base_operacional(empresa)
        except Exception:
            if caminho.exists():
                criar_backup_empresa(empresa, caminho)
            dados.to_csv(caminho, index=False, encoding="utf-8-sig")
            caminho_saida = caminho
    else:
        if caminho.exists():
            criar_backup_empresa(empresa, caminho)
        dados.to_csv(caminho, index=False, encoding="utf-8-sig")
        caminho_saida = caminho
    if st.session_state.get("salvamento_pendente", {}).get("empresa") == empresa:
        st.session_state.pop("salvamento_pendente", None)
    return caminho_saida


def formatar_moeda_br(valor: float) -> str:
    return f"R$ {valor:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")


def formatar_data_br(valor: object) -> str:
    data = pd.to_datetime(valor, errors="coerce")
    if pd.isna(data):
        return ""
    return data.strftime("%d/%m/%Y")


def campo_data_ptbr(
    container,
    label: str,
    value: date | datetime | None = None,
    key: str | None = None,
    placeholder: str = "dd/mm/aaaa",
):
    """Renderiza um seletor de data com calendario em pt-BR.

    Usa RSuite quando disponivel e cai para o widget nativo do Streamlit
    apenas como contingencia local.
    """
    valor_inicial = value
    if isinstance(valor_inicial, datetime):
        valor_inicial = valor_inicial.date()

    with container:
        if rsuite_date_picker is not None:
            return rsuite_date_picker(
                label=label,
                value=valor_inicial,
                format="dd/MM/yyyy",
                placeholder=placeholder,
                placement="bottomStart",
                one_tap=False,
                disabled=False,
                cleanable=True,
                block=True,
                iso_week=False,
                show_week_numbers=False,
                locale="pt_BR",
                key=key,
            )

        return st.date_input(
            label,
            value=valor_inicial or datetime.now().date(),
            format="DD/MM/YYYY",
            key=key,
        )


def agora_br() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def renderizar_botao_menu_lateral() -> None:
    components.html(
        """
        <button class="menu-lateral-fab" type="button" aria-label="Abrir menu lateral" title="Abrir menu lateral">
            ☰ Menu
        </button>
        <script>
            const btn = document.currentScript.previousElementSibling;
            btn.addEventListener("click", () => {
                const doc = window.parent?.document;
                if (!doc) return;
                const seletores = [
                    'button[data-testid="collapsedControl"]',
                    'button[aria-label*="sidebar" i]',
                    'button[title*="sidebar" i]'
                ];
                for (const seletor of seletores) {
                    const alvo = doc.querySelector(seletor);
                    if (alvo) {
                        alvo.click();
                        return;
                    }
                }
            });
        </script>
        <style>
            html, body {
                margin: 0;
                padding: 0;
                background: transparent;
                overflow: hidden;
            }
            .menu-lateral-fab {
                position: fixed;
                top: 0.85rem;
                left: 0.85rem;
                z-index: 9999;
                border: 1px solid #c8ddce;
                border-radius: 999px;
                padding: 0.55rem 0.85rem;
                background: rgba(255, 255, 255, 0.96);
                color: #173326;
                font-weight: 700;
                box-shadow: 0 8px 24px rgba(23, 51, 38, 0.12);
                cursor: pointer;
            }
            .menu-lateral-fab:hover {
                border-color: #2f8f5b;
                background: #ffffff;
            }
        </style>
        """,
        height=60,
    )


def aplicar_emojis_botoes_filtro() -> None:
    components.html(
        """
        <script>
            const emojis = new Map([
                ["Ativas", "🟢"],
                ["Excluídas", "🚫"],
                ["Excluidas", "🚫"],
                ["Importar", "📥"],
                ["Exportar", "📤"],
            ]);

            function ajustar() {
                const doc = window.parent?.document;
                if (!doc) return;

                doc.querySelectorAll("button").forEach((btn) => {
                    const texto = (btn.textContent || "").trim();
                    const emoji = emojis.get(texto);
                    if (!emoji || btn.dataset.emojiApplied === "1") return;
                    btn.dataset.emojiApplied = "1";
                    btn.textContent = `${emoji} ${texto}`;
                });
            }

            ajustar();
            const observer = new MutationObserver(ajustar);
            observer.observe(window.parent?.document?.body || document.body, { childList: true, subtree: true });
        </script>
        """,
        height=0,
    )


def voltar_menu_inicial() -> None:
    st.session_state.empresa_logada = ""
    st.session_state.modulo_atual = "contas_a_pagar"
    st.session_state.manutencao_ativa = ""
    st.session_state.exibir_troca_senha_padrao = False
    st.session_state.exibir_troca_senha_manual = False
    st.session_state.conta_edicao_indice = None
    st.session_state.conta_duplicacao_indice = None
    sincronizar_estado_url()
    st.rerun()


def renderizar_menu_direito(config: dict) -> None:
    _, col_inicio, col_menu = st.columns([7.8, 1.05, 1.15], gap="small")
    with col_inicio:
        st.button(
            "Inicio",
            key="btn_voltar_menu_inicial",
            help="Voltar para a tela inicial",
            use_container_width=True,
            type="secondary",
            on_click=voltar_menu_inicial,
        )

    with col_menu:
        if hasattr(st, "popover"):
            with st.popover("Menu", use_container_width=True):
                st.caption("Menu lateral")
                if st.session_state.get("empresa_logada"):
                    for modulo in MODULOS_CONTABEIS:
                        ativo = bool(modulo["ativo"])
                        selecionado = modulo["id"] == st.session_state.get("modulo_atual", "contas_a_pagar")
                        label = modulo["titulo"] if ativo else f"{modulo['titulo']} - em breve"
                        st.button(
                            label,
                            key=f"menu_direito_{modulo['id']}",
                            use_container_width=True,
                            disabled=not ativo,
                            type="primary" if selecionado and ativo else "secondary",
                            on_click=selecionar_modulo if ativo else None,
                            args=(modulo["id"],) if ativo else None,
                        )
                else:
                    st.caption("Selecione uma empresa para liberar os módulos.")
                st.divider()
                trocar_senha = globals().get("abrir_troca_senha")
                if trocar_senha:
                    st.button("Trocar senha", key="menu_direito_senha", use_container_width=True, on_click=trocar_senha)
                st.button("Sair", key="menu_direito_sair", use_container_width=True, on_click=logout)
        else:
            st.button("Menu", key="btn_menu_direito_fallback", use_container_width=True, type="secondary")


def gerar_csv_download(df: pd.DataFrame) -> bytes:
    colunas = [coluna for coluna in COLUNAS_DADOS if coluna in df.columns]
    return df[colunas].to_csv(index=False).encode("utf-8-sig")


def gerar_excel_contas_download(df: pd.DataFrame, empresa: str) -> bytes:
    dados = df.sort_values(["vencimento_dt", "descricao"], na_position="last").copy()
    hoje = pd.Timestamp(datetime.now().date())

    colunas = [
        ("vencimento", "Vencimento"),
        ("descricao", "Descricao"),
        ("fornecedor_cliente", "Fornecedor"),
        ("tipo_conta_nome", "Tipo de conta"),
        ("valor", "Valor"),
        ("status", "Status"),
        ("categoria", "Categoria"),
        ("documento", "Documento"),
        ("codigo_pagamento", "Codigo de barras / Pix"),
        ("observacao", "Observacao"),
        ("criado_em", "Incluido em"),
        ("criado_por", "Incluido por"),
    ]
    colunas = [(campo, titulo) for campo, titulo in colunas if campo in dados.columns]

    wb = Workbook()
    ws = wb.active
    ws.title = "Contas a pagar"
    ws.sheet_view.showGridLines = False

    cor_titulo = "173326"
    cor_verde = "2F8F5B"
    cor_verde_claro = "E9F5EC"
    cor_borda = "C8DDCE"
    cor_alerta = "FFF7E6"
    cor_vencida = "FEE2E2"
    cor_texto = "173326"

    ws.merge_cells("A1:L1")
    ws["A1"] = "Contas a pagar"
    ws["A1"].font = Font(bold=True, size=20, color=cor_titulo)
    ws["A1"].alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:L2")
    ws["A2"] = f"Empresa: {empresa} | Gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws["A2"].font = Font(size=10, color="60776A")

    total = float(pd.to_numeric(dados.get("valor", pd.Series(dtype=float)), errors="coerce").fillna(0).sum())
    vencidas = int(dados.apply(conta_vencida, axis=1).sum()) if not dados.empty else 0
    vence_hoje = int((dados.get("vencimento_dt", pd.Series(dtype="datetime64[ns]")).dt.date == hoje.date()).sum()) if "vencimento_dt" in dados else 0
    proximos_7 = 0
    if "vencimento_dt" in dados:
        proximos_7 = int(((dados["vencimento_dt"] > hoje) & (dados["vencimento_dt"] <= hoje + pd.Timedelta(days=7))).sum())

    resumo = [
        ("Total a pagar", total, "moeda"),
        ("Contas abertas", len(dados), "numero"),
        ("Vencidas", vencidas, "numero"),
        ("Vence hoje", vence_hoje, "numero"),
        ("Proximos 7 dias", proximos_7, "numero"),
    ]
    for indice, (rotulo, valor, tipo) in enumerate(resumo):
        coluna = 1 + (indice * 2)
        ws.merge_cells(start_row=4, start_column=coluna, end_row=4, end_column=coluna + 1)
        ws.merge_cells(start_row=5, start_column=coluna, end_row=5, end_column=coluna + 1)
        cel_rotulo = ws.cell(row=4, column=coluna, value=rotulo)
        cel_valor = ws.cell(row=5, column=coluna, value=valor)
        for linha in (4, 5):
            cel = ws.cell(row=linha, column=coluna)
            cel.fill = PatternFill("solid", fgColor=cor_verde_claro)
            cel.border = Border(
                left=Side(style="thin", color=cor_borda),
                right=Side(style="thin", color=cor_borda),
                top=Side(style="thin", color=cor_borda),
                bottom=Side(style="thin", color=cor_borda),
            )
            cel.alignment = Alignment(horizontal="center", vertical="center")
        cel_rotulo.font = Font(bold=True, size=9, color="60776A")
        cel_valor.font = Font(bold=True, size=14, color=cor_titulo)
        if tipo == "moeda":
            cel_valor.number_format = '"R$" #,##0.00'

    linha_cabecalho = 8
    for col_idx, (_, titulo) in enumerate(colunas, start=1):
        cel = ws.cell(row=linha_cabecalho, column=col_idx, value=titulo)
        cel.fill = PatternFill("solid", fgColor=cor_verde)
        cel.font = Font(bold=True, color="FFFFFF")
        cel.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cel.border = Border(bottom=Side(style="thin", color=cor_borda))

    for row_idx, (_, linha) in enumerate(dados.iterrows(), start=linha_cabecalho + 1):
        nivel = nivel_prazo_conta(linha)
        fill = None
        if nivel == "vencida":
            fill = PatternFill("solid", fgColor=cor_vencida)
        elif nivel == "atencao":
            fill = PatternFill("solid", fgColor=cor_alerta)

        for col_idx, (campo, _) in enumerate(colunas, start=1):
            valor = linha.get(campo, "")
            cel = ws.cell(row=row_idx, column=col_idx)
            if campo in {"vencimento", "criado_em"}:
                data = pd.to_datetime(valor, errors="coerce")
                cel.value = data.to_pydatetime() if not pd.isna(data) else ""
                cel.number_format = "dd/mm/yyyy"
            elif campo == "valor":
                cel.value = float(pd.to_numeric(valor, errors="coerce") or 0)
                cel.number_format = '"R$" #,##0.00'
            else:
                cel.value = str(valor or "")
            cel.font = Font(color=cor_texto)
            cel.alignment = Alignment(vertical="top", wrap_text=True)
            cel.border = Border(bottom=Side(style="hair", color=cor_borda))
            if fill:
                cel.fill = fill

    ultima_linha = max(linha_cabecalho + 1, linha_cabecalho + len(dados))
    ultima_coluna = max(1, len(colunas))
    tabela_ref = f"A{linha_cabecalho}:{get_column_letter(ultima_coluna)}{ultima_linha}"
    tabela = Table(displayName="TabelaContasPagar", ref=tabela_ref)
    tabela.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium4",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(tabela)

    larguras = {
        "Vencimento": 13,
        "Descricao": 32,
        "Fornecedor": 28,
        "Tipo de conta": 30,
        "Valor": 15,
        "Status": 14,
        "Categoria": 20,
        "Documento": 18,
        "Codigo de barras / Pix": 34,
        "Observacao": 34,
        "Incluido em": 16,
        "Incluido por": 14,
    }
    for col_idx, (_, titulo) in enumerate(colunas, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = larguras.get(titulo, 18)
    ws.freeze_panes = "A9"
    ws.auto_filter.ref = tabela_ref

    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    saida = BytesIO()
    wb.save(saida)
    return saida.getvalue()


def opcoes_com_historico(df: pd.DataFrame, coluna: str, padroes: list[str]) -> list[str]:
    opcoes = [""]
    if coluna in df.columns:
        historico = (
            df[coluna]
            .dropna()
            .astype(str)
            .map(str.strip)
            .loc[lambda serie: serie.ne("")]
            .drop_duplicates()
            .tolist()
        )
        opcoes.extend(historico)
    opcoes.extend(padroes)
    unicas = []
    chaves_vistas = set()
    for opcao in opcoes:
        texto = str(opcao or "").strip()
        chave = unicodedata.normalize("NFKD", texto).encode("ascii", "ignore").decode("ascii").casefold()
        chave = re.sub(r"\s+", " ", chave).strip()
        if chave in OPCOES_OCULTAS:
            continue
        if chave in chaves_vistas:
            continue
        chaves_vistas.add(chave)
        unicas.append(texto)
    return unicas


def resolver_opcao_digitavel(selecao: str, texto_novo: str) -> str:
    if selecao == OPCAO_NOVO:
        return texto_novo.strip()
    return str(selecao or "").strip()


def opcoes_com_novo(opcoes: list[str]) -> list[str]:
    limpas = [
        str(opcao or "").strip()
        for opcao in opcoes
        if str(opcao or "").strip() and str(opcao or "").strip() != OPCAO_NOVO
    ]
    ordenadas = sorted(
        limpas,
        key=lambda texto: unicodedata.normalize("NFKD", texto).encode("ascii", "ignore").decode("ascii").casefold(),
    )
    return [OPCAO_NOVO, *ordenadas]


def rotulo_opcao_nova(opcao: str) -> str:
    return f"🔴 {OPCAO_NOVO}" if opcao == OPCAO_NOVO else opcao


def indice_padrao_sem_novo(opcoes: list[str]) -> int:
    for indice, opcao in enumerate(opcoes):
        if str(opcao or "").strip() and opcao != OPCAO_NOVO:
            return indice
    return 0


def campo_opcao_nova(
    container: object,
    label: str,
    opcoes: list[str],
    label_novo: str,
    key_prefix: str,
    index: int | None = None,
) -> tuple[str, str]:
    indice = indice_padrao_sem_novo(opcoes) if index is None else index
    selecao = container.selectbox(
        label,
        opcoes,
        index=indice,
        key=f"{key_prefix}_selecao",
        format_func=rotulo_opcao_nova,
    )
    texto_novo = ""
    if selecao == OPCAO_NOVO:
        texto_novo = container.text_input(label_novo, key=f"{key_prefix}_novo")
    return selecao, texto_novo


def gerar_modelo_importacao_csv() -> bytes:
    exemplo = pd.DataFrame(
        [
            {
                "descricao": "Boleto fornecedor exemplo",
                "fornecedor": "Fornecedor Exemplo",
                "vencimento": "2026-06-10",
                "valor": "1500,00",
                "status": "aberto",
                "categoria": "Administrativo",
                "documento": "NF-0001",
                "tipo_conta_codigo": "2.1.6.02.001",
                "codigo_pagamento": "34191.79001 01043.510047 91020.150008 8 123400000150000",
                "observacao": "Linha de exemplo; apague antes de importar dados reais.",
            }
        ],
        columns=COLUNAS_IMPORTACAO,
    )
    return exemplo.to_csv(index=False, sep=";", encoding="utf-8-sig").encode("utf-8-sig")


def ler_csv_importacao(arquivo: object) -> pd.DataFrame:
    try:
        return pd.read_csv(arquivo, sep=None, engine="python", dtype=str).fillna("")
    except Exception as erro:
        raise ValueError(f"Não foi possível ler o CSV: {erro}") from erro


def normalizar_valor_importado(valor: object) -> float:
    texto = str(valor).strip()
    if not texto:
        return 0.0
    texto = re.sub(r"[^0-9,.-]", "", texto)
    if "," in texto and "." in texto:
        texto = texto.replace(".", "").replace(",", ".")
    elif "," in texto:
        texto = texto.replace(",", ".")
    try:
        return float(texto)
    except ValueError:
        return 0.0


def parse_valor_br(valor: object) -> float | None:
    texto = str(valor or "").strip()
    if not texto:
        return None

    texto = re.sub(r"[^0-9,.-]", "", texto)
    if not texto:
        return None
    if "," in texto:
        texto = texto.replace(".", "").replace(",", ".")
    elif "." in texto:
        texto = texto.replace(".", "")

    try:
        return float(texto)
    except ValueError:
        return None


def parse_data_br(valor: object) -> pd.Timestamp | None:
    texto = str(valor or "").strip()
    if not texto:
        return None
    data = pd.to_datetime(texto, errors="coerce", dayfirst=True)
    if pd.isna(data):
        return None
    return pd.Timestamp(data)


def preparar_importacao_contas(df_importado: pd.DataFrame, empresa: str, usuario: str) -> tuple[pd.DataFrame, list[str]]:
    df_importado.columns = [str(coluna).strip().lower() for coluna in df_importado.columns]
    faltantes = [coluna for coluna in COLUNAS_IMPORTACAO_OBRIGATORIAS if coluna not in df_importado.columns]
    if faltantes:
        raise ValueError(f"Colunas obrigatórias ausentes: {', '.join(faltantes)}.")

    linhas = []
    erros = []
    for numero_linha, linha in df_importado.iterrows():
        linha_numero = numero_linha + 2
        descricao = str(linha.get("descricao", "")).strip()
        fornecedor = str(linha.get("fornecedor", "")).strip()
        vencimento = pd.to_datetime(linha.get("vencimento", ""), errors="coerce", dayfirst=True)
        valor = normalizar_valor_importado(linha.get("valor", ""))
        status = str(linha.get("status", "aberto")).strip().lower() or "aberto"
        status = status if status in {"aberto", "pendente", "vencido"} else "aberto"
        tipo_codigo = str(linha.get("tipo_conta_codigo", "")).strip()
        tipo_nome = TIPOS_CONTA_PAGAR.get(tipo_codigo, "")
        codigo_pagamento = str(linha.get("codigo_pagamento", "")).strip()

        if not descricao or not fornecedor or pd.isna(vencimento) or valor <= 0:
            erros.append(f"Linha {linha_numero}: preencha descricao, fornecedor, vencimento válido e valor maior que zero.")
            continue

        documento = str(linha.get("documento", "")).strip()
        if not documento:
            documento = f"IMP-{empresa}-{datetime.now().strftime('%Y%m%d%H%M%S')}-{linha_numero}"

        linhas.append(
            {
                "empresa": empresa,
                "competencia": "",
                "tipo": "conta_a_pagar",
                "descricao": descricao,
                "fornecedor_cliente": fornecedor,
                "vencimento": vencimento.strftime("%Y-%m-%d"),
                "pagamento_recebimento": "",
                "valor": valor,
                "status": status,
                "categoria": str(linha.get("categoria", "")).strip(),
                "observacao": str(linha.get("observacao", "")).strip(),
                "documento": documento,
                "tipo_conta_codigo": tipo_codigo,
                "tipo_conta_nome": tipo_nome,
                "anexo_nome": "",
                "anexo_caminho": "",
                "codigo_pagamento": codigo_pagamento,
                "criado_em": agora_br(),
                "criado_por": usuario,
                "excluido_em": "",
                "excluido_por": "",
                "ativo": True,
            }
        )

    return pd.DataFrame(linhas, columns=COLUNAS_DADOS), erros


def validar_login(config: dict, usuario: str, senha_digitada: str) -> bool:
    senha_configurada = obter_senha_usuario(config, usuario)
    return bool(senha_configurada and senha_digitada and senha_digitada == senha_configurada)


def obter_cliente(config: dict, empresa: str) -> dict:
    return next((item for item in config["clientes"] if item["empresa"] == empresa), {})


def obter_usuario(config: dict, usuario: str) -> dict:
    usuario_normalizado = usuario.strip().upper()
    return next(
        (item for item in config["usuarios"] if item.get("usuario", "").strip().upper() == usuario_normalizado),
        {},
    )


def obter_senha_usuario(config: dict, usuario: str) -> str:
    usuario_config = obter_usuario(config, usuario)
    usuario_chave = usuario_config.get("usuario", usuario).strip().upper()
    try:
        usuarios_secret = st.secrets.get("usuarios", {})
        if usuario_chave in usuarios_secret:
            return str(usuarios_secret[usuario_chave])
    except Exception:
        pass

    return str(usuario_config.get("senha", ""))


def empresas_do_usuario(config: dict, usuario: str) -> list[str]:
    usuario_config = obter_usuario(config, usuario)
    empresas_liberadas = usuario_config.get("empresas", [])
    empresas_configuradas = [cliente["empresa"] for cliente in config["clientes"]]
    if not empresas_liberadas:
        return empresas_configuradas
    return [empresa for empresa in empresas_configuradas if empresa in empresas_liberadas]


def inicializar_sessao() -> None:
    st.session_state.setdefault("autenticado", False)
    st.session_state.setdefault("auth_token", "")
    st.session_state.setdefault("usuario_logado", "")
    st.session_state.setdefault("empresa_logada", "")
    st.session_state.setdefault("modulo_atual", "contas_a_pagar")
    st.session_state.setdefault("contas_visualizacao_modo", "Tabela")
    st.session_state.setdefault("exibir_troca_senha_padrao", False)
    st.session_state.setdefault("senha_padrao_mantida", False)
    st.session_state.setdefault("exibir_troca_senha_manual", False)


def ler_estado_url() -> dict[str, str]:
    try:
        params = st.query_params
        return {chave: str(valor) for chave, valor in params.items()}
    except Exception:
        params = st.experimental_get_query_params()
        return {chave: str(valor[0]) for chave, valor in params.items() if valor}


def definir_estado_url(**estado: object) -> None:
    dados = {chave: str(valor) for chave, valor in estado.items() if valor is not None and str(valor) != ""}
    try:
        if hasattr(st, "query_params"):
            st.query_params.clear()
            for chave, valor in dados.items():
                st.query_params[chave] = valor
        else:
            st.experimental_set_query_params(**dados)
    except Exception:
        pass


def limpar_estado_url() -> None:
    try:
        if hasattr(st, "query_params"):
            st.query_params.clear()
        else:
            st.experimental_set_query_params()
    except Exception:
        pass


def carregar_sessoes_autenticacao() -> dict[str, dict[str, object]]:
    if not SESSOES_PATH.exists():
        return {}
    try:
        dados = json.loads(SESSOES_PATH.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, OSError):
        return {}
    return dados if isinstance(dados, dict) else {}


def salvar_sessoes_autenticacao(sessoes: dict[str, dict[str, object]]) -> None:
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    SESSOES_PATH.write_text(json.dumps(sessoes, indent=2, ensure_ascii=False), encoding="utf-8")


def criar_sessao_autenticacao(usuario: str) -> str:
    agora = time.time()
    sessoes = {
        token: dados
        for token, dados in carregar_sessoes_autenticacao().items()
        if float(dados.get("expira_em", 0) or 0) > agora
    }
    token = secrets.token_urlsafe(32)
    sessoes[token] = {
        "usuario": usuario.strip().upper(),
        "expira_em": agora + SESSAO_LOGIN_SEGUNDOS,
    }
    salvar_sessoes_autenticacao(sessoes)
    return token


def obter_sessao_autenticacao(token: str) -> dict[str, object] | None:
    if not token:
        return None
    sessoes = carregar_sessoes_autenticacao()
    sessao = sessoes.get(token)
    if not sessao:
        return None
    if float(sessao.get("expira_em", 0) or 0) <= time.time():
        sessoes.pop(token, None)
        salvar_sessoes_autenticacao(sessoes)
        return None
    return sessao


def invalidar_sessao_autenticacao(token: str) -> None:
    if not token:
        return
    sessoes = carregar_sessoes_autenticacao()
    if token in sessoes:
        sessoes.pop(token, None)
        salvar_sessoes_autenticacao(sessoes)


def expirar_sessao_se_necessario() -> None:
    if not st.session_state.get("autenticado"):
        return
    token = str(st.session_state.get("auth_token", ""))
    if obter_sessao_autenticacao(token):
        return
    st.session_state.autenticado = False
    st.session_state.auth_token = ""
    st.session_state.usuario_logado = ""
    st.session_state.empresa_logada = ""
    st.session_state.exibir_troca_senha_padrao = False
    st.session_state.exibir_troca_senha_manual = False
    limpar_estado_url()


def restaurar_estado_sessao() -> None:
    estado = ler_estado_url()
    sessao = obter_sessao_autenticacao(estado.get("token", ""))
    if not sessao:
        if estado.get("auth") == "1" or estado.get("token"):
            limpar_estado_url()
        return

    st.session_state.autenticado = True
    st.session_state.auth_token = estado.get("token", "")
    st.session_state.usuario_logado = str(sessao.get("usuario", ""))
    st.session_state.empresa_logada = estado.get("empresa", st.session_state.get("empresa_logada", ""))
    st.session_state.modulo_atual = estado.get("modulo", st.session_state.get("modulo_atual", "contas_a_pagar"))
    st.session_state.exibir_troca_senha_padrao = estado.get("troca_padrao") == "1"
    st.session_state.exibir_troca_senha_manual = estado.get("troca_manual") == "1"
    st.session_state.senha_padrao_mantida = estado.get("senha_padrao_mantida") == "1"


def sincronizar_estado_url() -> None:
    if not st.session_state.get("autenticado"):
        limpar_estado_url()
        return
    token = str(st.session_state.get("auth_token", ""))
    if token and not obter_sessao_autenticacao(token):
        st.session_state.autenticado = False
        st.session_state.auth_token = ""
        limpar_estado_url()
        st.rerun()
    if not token:
        token = criar_sessao_autenticacao(str(st.session_state.get("usuario_logado", "")))
        st.session_state.auth_token = token

    definir_estado_url(
        auth="1",
        token=token,
        usuario=st.session_state.get("usuario_logado", ""),
        empresa=st.session_state.get("empresa_logada", ""),
        modulo=st.session_state.get("modulo_atual", "contas_a_pagar"),
        troca_padrao="1" if st.session_state.get("exibir_troca_senha_padrao") else "",
        troca_manual="1" if st.session_state.get("exibir_troca_senha_manual") else "",
        senha_padrao_mantida="1" if st.session_state.get("senha_padrao_mantida") else "",
    )


def tela_login(config: dict) -> None:
    logo_b64 = imagem_base64(LOGO_FULL_PATH)
    logo_html = (
        f'<img src="data:image/png;base64,{logo_b64}" alt="MH LOG">'
        if logo_b64
        else '<strong style="color:var(--mh-accent);font-size:1.5rem;">MH LOG</strong>'
    )

    st.write("")
    _, col_login, _ = st.columns([1, 1.05, 1])
    with col_login:
        st.markdown(
            f"""
            <div class="login-logo">{logo_html}</div>
            <section class="login-panel">
                <span class="login-badge">Acesso restrito</span>
                <h2>Portal Contábil do Cliente</h2>
                <p>Contas a pagar liberadas pelo escritório.</p>
                <p>Use seu usuário e senha de acesso.</p>
            </section>
            """,
            unsafe_allow_html=True,
        )
        st.write("")
        with st.container(border=True):
            st.subheader("Entrar no portal")
            st.caption("Informe seus dados para continuar.")
            with st.form("form_login", clear_on_submit=False):
                empresa = st.selectbox("Usuário", empresas, index=empresa_idx)
                senha = st.text_input("Senha de acesso", type="password")
                entrar = st.form_submit_button("Entrar", use_container_width=True)

            if entrar:
                cliente_atual = obter_cliente(config, empresa)
                competencias = cliente_atual.get("competencias", [])
                if validar_login(config, empresa, senha):
                    st.session_state.autenticado = True
                    st.session_state.empresa_logada = empresa
                    st.session_state.competencia_logada = competencias[0] if competencias else ""
                    st.session_state.empresa_login = empresa
                    st.rerun()
                else:
                    st.error("Usuário ou senha inválida.")


def logout() -> None:
    st.session_state.autenticado = False
    st.session_state.empresa_logada = ""
    st.session_state.competencia_logada = ""
    st.rerun()


def sidebar_filtros_antigo(config: dict) -> tuple[str, str]:
    if LOGO_FULL_PATH.exists():
        st.sidebar.image(str(LOGO_FULL_PATH), width=210)
    st.sidebar.title("🔐 Sessão do Cliente")

    empresa = st.session_state.get("empresa_logada", "")
    st.sidebar.text_input("Usuário", value=empresa, disabled=True)
    st.sidebar.divider()
    st.sidebar.success("Acesso liberado")
    st.sidebar.button("Sair", use_container_width=True, on_click=logout)
    return empresa, ""


def filtrar_dados(df: pd.DataFrame, empresa: str) -> pd.DataFrame:
    filtro = df["empresa"] == empresa
    return remover_linhas_em_branco(df.loc[filtro].copy())


def filtrar_contas_a_pagar_abertas(df: pd.DataFrame, empresa: str) -> pd.DataFrame:
    dados = filtrar_dados(df, empresa)
    status_abertos = dados["status"].astype(str).str.lower().isin(["aberto", "pendente", "vencido"])
    filtro = (dados["tipo"] == "conta_a_pagar") & status_abertos & dados["ativo"]
    return remover_linhas_em_branco(dados.loc[filtro].copy())


def tela_login(config: dict) -> None:
    logo_b64 = imagem_base64(LOGO_FULL_PATH)
    logo_html = (
        f'<img src="data:image/png;base64,{logo_b64}" alt="MH LOG">'
        if logo_b64
        else '<strong style="color:var(--mh-accent);font-size:1.5rem;">MH LOG</strong>'
    )

    st.write("")
    col_vazio_esq, col_esquerda, col_direita, col_vazio_dir = st.columns([0.65, 2.35, 2.35, 0.65], gap="small")

    with col_esquerda:
        st.markdown(
            f"""
            <div class="login-logo">{logo_html}</div>
            <section class="login-panel">
                <span class="login-badge">Acesso restrito</span>
                <h2>Portal Contabil do Cliente</h2>
                <p>Contas a pagar liberadas pelo escritorio.</p>
                <p>Use seu usuario e senha de acesso.</p>
            </section>
            """,
            unsafe_allow_html=True,
        )

    with col_direita:
        with st.container(border=True):
            st.subheader("Entrar no portal")
            st.caption("Informe seus dados para continuar.")
            with st.form("form_login", clear_on_submit=False):
                usuario = st.text_input("Usuario", value=st.session_state.get("usuario_login", ""))
                senha = st.text_input("Senha de acesso", type="password")
                entrar = st.form_submit_button("Entrar", use_container_width=True)

            if entrar:
                usuario = usuario.strip().upper()
                empresas_liberadas = empresas_do_usuario(config, usuario)
                if validar_login(config, usuario, senha) and empresas_liberadas:
                    st.session_state.autenticado = True
                    st.session_state.usuario_logado = usuario
                    st.session_state.usuario_login = usuario
                    st.session_state.empresa_logada = ""
                    st.session_state.modulo_atual = "contas_a_pagar"
                    st.session_state.exibir_troca_senha_padrao = senha == "123456"
                    st.session_state.exibir_troca_senha_manual = False
                    st.session_state.senha_padrao_mantida = False
                    sincronizar_estado_url()
                    st.rerun()
                else:
                    st.error("Usuario ou senha invalida.")


def logout() -> None:
    invalidar_sessao_autenticacao(str(st.session_state.get("auth_token", "")))
    st.session_state.autenticado = False
    st.session_state.auth_token = ""
    st.session_state.usuario_logado = ""
    st.session_state.empresa_logada = ""
    st.session_state.exibir_troca_senha_padrao = False
    st.session_state.senha_padrao_mantida = False
    st.session_state.exibir_troca_senha_manual = False
    limpar_estado_url()


def voltar_dashboard_empresas() -> None:
    st.session_state.empresa_logada = ""
    st.session_state.modulo_atual = "contas_a_pagar"
    sincronizar_estado_url()


def selecionar_empresa(empresa: str, config: dict) -> None:
    st.session_state.empresa_logada = empresa
    st.session_state.modulo_atual = "contas_a_pagar"
    sincronizar_estado_url()


def selecionar_modulo(modulo_id: str) -> None:
    st.session_state.modulo_atual = modulo_id
    sincronizar_estado_url()


def exibir_menu_contabil() -> None:
    st.sidebar.markdown(
        """
        <div class="menu-contabil-titulo">Menu contábil</div>
        <div class="menu-contabil-ajuda">Os módulos desativados serão liberados aos poucos.</div>
        """,
        unsafe_allow_html=True,
    )
    modulo_atual = st.session_state.get("modulo_atual", "contas_a_pagar")
    for modulo in MODULOS_CONTABEIS:
        ativo = bool(modulo["ativo"])
        selecionado = modulo["id"] == modulo_atual
        label = modulo["titulo"] if ativo else f"{modulo['titulo']} - em breve"
        st.sidebar.button(
            label,
            key=f"menu_{modulo['id']}",
            use_container_width=True,
            disabled=not ativo,
            type="primary" if selecionado and ativo else "secondary",
            on_click=selecionar_modulo if ativo else None,
            args=(modulo["id"],) if ativo else None,
        )


def sidebar_filtros(config: dict) -> tuple[str, str]:
    if LOGO_FULL_PATH.exists():
        st.sidebar.image(str(LOGO_FULL_PATH), width=120)

    usuario = st.session_state.get("usuario_logado", "")
    empresa = st.session_state.get("empresa_logada", "")

    exibir_menu_contabil()
    st.sidebar.divider()
    st.sidebar.title("Sessao")
    st.sidebar.text_input("Usuario", value=usuario, disabled=True)
    st.sidebar.text_input("Empresa", value=empresa, disabled=True)
    st.sidebar.divider()
    st.sidebar.success("Acesso liberado")
    st.sidebar.button("Trocar empresa", use_container_width=True, on_click=voltar_dashboard_empresas)
    st.sidebar.button("Sair", use_container_width=True, on_click=logout)
    return empresa, usuario


def dashboard_empresas(config: dict, df: pd.DataFrame) -> None:
    usuario = st.session_state.get("usuario_logado", "")
    empresas = empresas_do_usuario(config, usuario)

    if LOGO_FULL_PATH.exists():
        st.sidebar.image(str(LOGO_FULL_PATH), width=120)
    st.sidebar.markdown(
        """
        <div class="menu-contabil-titulo">Menu contábil</div>
        <div class="menu-contabil-ajuda">Escolha uma empresa para liberar os módulos.</div>
        """,
        unsafe_allow_html=True,
    )
    for modulo in MODULOS_CONTABEIS:
        label = modulo["titulo"] if modulo["ativo"] else f"{modulo['titulo']} - em breve"
        st.sidebar.button(label, key=f"dashboard_menu_{modulo['id']}", use_container_width=True, disabled=True)
    st.sidebar.divider()
    st.sidebar.title("Sessao")
    st.sidebar.text_input("Usuario", value=usuario, disabled=True)
    st.sidebar.button("Sair", use_container_width=True, on_click=logout)

    left, center, right = st.columns([0.8, 10.4, 0.8], gap="small")
    with center:
        st.title("Dashboard por empresa")
        st.caption("Escolha uma empresa para abrir as contas a pagar e os indicadores dela.")

        cols = st.columns(2, gap="large")
        for indice, empresa in enumerate(empresas):
            contas_empresa = df.loc[
                (df["empresa"] == empresa)
                & (df["tipo"] == "conta_a_pagar")
                & df["status"].astype(str).str.lower().isin(["aberto", "pendente", "vencido"])
                & df["ativo"]
            ]
            total = contas_empresa["valor"].sum()
            abertas = int(len(contas_empresa))
            vencidas = int(contas_empresa.apply(conta_vencida, axis=1).sum())
            vence_hoje = int((contas_empresa["vencimento_dt"].dt.date == datetime.now().date()).sum())
            status_classe = "vencidas" if vencidas else ("abertas" if abertas else "info")
            status_texto = "Prioridade alta" if vencidas else ("Em acompanhamento" if abertas else "Sem pendências")

            with cols[indice % 2]:
                st.markdown(
                    f"""
                    <div class="empresa-card">
                        <div class="empresa-card-topo">
                            <div>
                                <h3>{escape(empresa)}</h3>
                                <div style="color: var(--mh-muted); font-size: 0.86rem;">Painel operacional da empresa</div>
                            </div>
                            <div class="empresa-chip {status_classe}">{"🔴" if vencidas else ("🟢" if abertas else "ℹ️")} {escape(status_texto)}</div>
                        </div>
                        <div class="empresa-card-badges">
                            <span class="empresa-chip total">💰 {escape(formatar_moeda_br(total))}</span>
                            <span class="empresa-chip abertas">📄 {abertas} contas</span>
                            <span class="empresa-chip vencidas">⏰ {vencidas} vencidas</span>
                            <span class="empresa-chip info">📅 {vence_hoje} vence hoje</span>
                        </div>
                        <p>Abra a empresa para acompanhar pagamentos, editar contas e registrar baixas.</p>
                    </div>
                    """,
                    unsafe_allow_html=True,
                )
                st.button(
                    f"Abrir {empresa}",
                    key=f"abrir_empresa_{empresa}",
                    use_container_width=True,
                    on_click=selecionar_empresa,
                    args=(empresa, config),
                )


def status_geral_contas(df: pd.DataFrame) -> tuple[str, str]:
    if df.empty:
        return "Sem contas em aberto", "ok"

    vencidos = int(df.apply(conta_vencida, axis=1).sum())
    if vencidos:
        return f"{vencidos} conta(s) vencida(s)", "alerta"
    return f"{len(df)} conta(s) em aberto", "pendente"


def conta_vencida(linha: pd.Series) -> bool:
    status = str(linha.get("status", linha.get("Status", ""))).strip().lower()
    if status == "vencido":
        return True
    if status not in {"aberto", "pendente"}:
        return False

    vencimento_valor = linha.get("vencimento", linha.get("Vencimento", ""))
    vencimento_texto = str(vencimento_valor)
    vencimento = pd.to_datetime(vencimento_valor, errors="coerce", dayfirst="/" in vencimento_texto)
    if pd.isna(vencimento):
        return False
    return vencimento.date() < datetime.now().date()


def nivel_prazo_conta(linha: pd.Series) -> str:
    if conta_vencida(linha):
        return "vencida"

    status = str(linha.get("status", linha.get("Status", ""))).strip().lower()
    if status not in {"aberto", "pendente"}:
        return "neutra"

    vencimento_valor = linha.get("vencimento", linha.get("Vencimento", ""))
    vencimento_texto = str(vencimento_valor)
    vencimento = pd.to_datetime(vencimento_valor, errors="coerce", dayfirst="/" in vencimento_texto)
    if pd.isna(vencimento):
        return "neutra"

    dias = (vencimento.date() - datetime.now().date()).days
    if dias <= 7:
        return "atencao"
    return "ok"


def indicador_prazo_conta(linha: pd.Series) -> str:
    nivel = nivel_prazo_conta(linha)
    if nivel == "vencida":
        return "🔴"
    if nivel == "atencao":
        return "🟡"
    if nivel == "ok":
        return "🟢"
    return "⚪"


def escrever_celula_conta(coluna: object, texto: str, nivel: str, nowrap: bool = False, indicador: str = "") -> None:
    cor = {"vencida": "#b91c1c", "atencao": "#b7791f", "ok": "inherit"}.get(nivel, "inherit")
    peso = "700" if nivel == "vencida" else "400"
    quebra = "white-space:nowrap;" if nowrap else "overflow-wrap:anywhere;"
    conteudo = f"{indicador} {texto}".strip()
    coluna.markdown(
        f'<span style="color:{cor};font-weight:{peso};{quebra}">{escape(conteudo)}</span>',
        unsafe_allow_html=True,
    )


def preparar_tabela_contas(df: pd.DataFrame) -> pd.DataFrame:
    colunas = [
        "vencimento",
        "descricao",
        "fornecedor_cliente",
        "tipo_conta_nome",
        "valor",
        "status",
        "categoria",
        "documento",
        "anexo_nome",
        "codigo_pagamento",
        "criado_em",
        "criado_por",
    ]
    dados = remover_linhas_em_branco(df)
    dados = dados.sort_values(["vencimento_dt", "descricao"], na_position="last").copy()
    dados = dados[[coluna for coluna in colunas if coluna in dados.columns]]
    if "vencimento" in dados:
        dados["vencimento"] = dados["vencimento"].apply(formatar_data_br)
    if "valor" in dados:
        dados["valor"] = dados["valor"].apply(formatar_moeda_br)
    return dados.rename(
        columns={
            "vencimento": "Vencimento",
            "descricao": "Descrição",
            "fornecedor_cliente": "Fornecedor",
            "tipo_conta_nome": "Tipo de conta",
            "valor": "Valor",
            "status": "Status",
            "categoria": "Categoria",
            "documento": "Documento",
            "anexo_nome": "Anexo",
            "codigo_pagamento": "Codigo de barras / Pix",
            "criado_em": "Incluído em",
            "criado_por": "Incluído por",
        }
    )


def altura_dataframe(qtd_linhas: int) -> int:
    if qtd_linhas <= 0:
        return 120
    return 46 + (qtd_linhas * 36)


def estilo_status_linha(linha: pd.Series) -> list[str]:
    status = str(linha.get("Status", "")).strip().lower()
    if conta_vencida(linha):
        estilo = "background-color: rgba(220, 38, 38, 0.10); color: #7f1d1d;"
    elif status in {"aberto", "pendente"}:
        estilo = "background-color: rgba(183, 121, 31, 0.08); color: #713f12;"
    else:
        estilo = ""
    return [estilo] * len(linha)


def exibir_contas(df: pd.DataFrame) -> None:
    tabela = preparar_tabela_contas(df)
    st.dataframe(
        tabela.style.apply(estilo_status_linha, axis=1),
        height=altura_dataframe(len(tabela)),
        use_container_width=True,
        hide_index=True,
    )


ORDENACOES_CONTAS = {
    "vencimento": {"label": "Vencimento", "coluna": "vencimento_dt", "tipo": "data"},
    "descricao": {"label": "Descricao", "coluna": "descricao", "tipo": "texto"},
    "fornecedor": {"label": "Fornecedor", "coluna": "fornecedor_cliente", "tipo": "texto"},
    "valor": {"label": "Valor", "coluna": "valor", "tipo": "numero"},
    "status": {"label": "Status", "coluna": "status", "tipo": "texto"},
    "documento": {"label": "Documento", "coluna": "documento", "tipo": "texto"},
}


def selecionar_ordenacao_contas(campo: str) -> None:
    atual = st.session_state.get("contas_ordem_campo", "")
    crescente = st.session_state.get("contas_ordem_crescente", True)
    st.session_state.contas_ordem_campo = campo
    if atual == campo:
        st.session_state.contas_ordem_crescente = not crescente
    else:
        st.session_state.contas_ordem_crescente = False if campo == "vencimento" else True


def rotulo_ordenacao(campo: str) -> str:
    info = ORDENACOES_CONTAS[campo]
    if st.session_state.get("contas_ordem_campo") != campo:
        return f"{info['label']} ↕"
    return f"{info['label']} {'↑' if st.session_state.get('contas_ordem_crescente', True) else '↓'}"


def ordenar_contas_exibidas(contas: pd.DataFrame) -> pd.DataFrame:
    campo = st.session_state.get("contas_ordem_campo", "")
    if campo not in ORDENACOES_CONTAS:
        dados = contas.copy()
        dados["_ordem_vencimento"] = pd.to_datetime(dados.get("vencimento_dt", dados.get("vencimento", "")), errors="coerce")
        dados["_ordem_recente"] = pd.to_datetime(dados.get("criado_em", ""), errors="coerce")
        dados["_ordem_documento"] = dados.get("documento", "").astype(str)
        return dados.sort_values(
            ["_ordem_vencimento", "_ordem_recente", "_ordem_documento"],
            ascending=[False, False, False],
            na_position="last",
        ).drop(columns=["_ordem_vencimento", "_ordem_recente", "_ordem_documento"], errors="ignore")

    info = ORDENACOES_CONTAS[campo]
    coluna = str(info["coluna"])
    crescente = bool(st.session_state.get("contas_ordem_crescente", True))
    dados = contas.copy()

    if info["tipo"] == "data":
        dados["_ordem"] = pd.to_datetime(dados.get(coluna, ""), errors="coerce")
    elif info["tipo"] == "numero":
        dados["_ordem"] = pd.to_numeric(dados.get(coluna, 0), errors="coerce").fillna(0)
    else:
        dados["_ordem"] = (
            dados.get(coluna, "")
            .fillna("")
            .astype(str)
            .map(lambda texto: unicodedata.normalize("NFKD", texto).encode("ascii", "ignore").decode("ascii").casefold())
        )

    return dados.sort_values("_ordem", ascending=crescente, na_position="last").drop(columns=["_ordem"], errors="ignore")


def exibir_contas_com_acoes(contas_exibidas: pd.DataFrame, df_base: pd.DataFrame, empresa: str, usuario: str) -> None:
    if contas_exibidas.empty:
        st.info("Nenhum registro encontrado para este filtro.")
        return

    contas = ordenar_contas_exibidas(contas_exibidas)
    larguras = [1.2, 2.2, 2.05, 1.1, 0.9, 1.15, 1.2]
    cabecalho = st.columns(larguras)
    for col, campo in zip(cabecalho[:6], ORDENACOES_CONTAS.keys()):
        col.button(
            rotulo_ordenacao(campo),
            key=f"ordenar_contas_{campo}",
            on_click=selecionar_ordenacao_contas,
            args=(campo,),
            use_container_width=True,
            type="secondary",
        )
    cabecalho[6].caption("Ações")

    for indice, linha in contas.iterrows():
        nivel = nivel_prazo_conta(linha)
        indicador = indicador_prazo_conta(linha)
        cols = st.columns(larguras)
        escrever_celula_conta(cols[0], formatar_data_br(linha.get("vencimento")), nivel, nowrap=True, indicador=indicador)
        escrever_celula_conta(cols[1], str(linha.get("descricao", "")), nivel)
        escrever_celula_conta(cols[2], str(linha.get("fornecedor_cliente", "")), nivel)
        escrever_celula_conta(cols[3], formatar_moeda_br(float(linha.get("valor", 0) or 0)), nivel, nowrap=True)
        escrever_celula_conta(cols[4], "vencido" if nivel == "vencida" else str(linha.get("status", "")), nivel, nowrap=True, indicador=indicador)
        escrever_celula_conta(cols[5], str(linha.get("documento", "")), nivel)
        acao_cols = cols[6].columns([1, 1, 1, 1], gap="small")
        if acao_cols[0].button("✏️", key=f"contas_acao_editar_{indice}", help="Editar conta", type="secondary", use_container_width=True):
            selecionar_conta_para_edicao(indice)
            st.rerun()
        if acao_cols[1].button("📋", key=f"contas_acao_duplicar_{indice}", help="Duplicar conta", type="secondary", use_container_width=True):
            selecionar_conta_para_duplicacao(indice)
            st.rerun()
        if acao_cols[2].button("✅", key=f"contas_acao_pagar_{indice}", help="Marcar como pago", type="secondary", use_container_width=True):
            try:
                df_atualizado = marcar_conta_como_paga(df_base, indice, usuario)
                if salvar_dados_empresa(normalizar_dataframe(df_atualizado), empresa) is None:
                    return
                st.success("Conta marcada como paga.")
                st.rerun()
            except ValueError as erro:
                st.error(str(erro))
        if acao_cols[3].button("🗑️", key=f"contas_acao_excluir_{indice}", help="Excluir conta", type="secondary", use_container_width=True):
            try:
                df_atualizado = excluir_conta_a_pagar(df_base, indice, usuario)
                if salvar_dados_empresa(normalizar_dataframe(df_atualizado), empresa) is None:
                    return
                st.success("Conta excluída da lista ativa.")
                st.rerun()
            except ValueError as erro:
                st.error(str(erro))


def _renderizar_acoes_conta(
    container: object,
    indice: int,
    df_base: pd.DataFrame,
    empresa: str,
    usuario: str,
    prefixo: str,
) -> None:
    botoes = container.columns([1, 1, 1, 1], gap="small")
    if botoes[0].button("✏️", key=f"{prefixo}_editar_{indice}", help="Editar conta", type="secondary", use_container_width=True):
        selecionar_conta_para_edicao(indice)
        st.rerun()
    if botoes[1].button("📋", key=f"{prefixo}_duplicar_{indice}", help="Duplicar conta", type="secondary", use_container_width=True):
        selecionar_conta_para_duplicacao(indice)
        st.rerun()
    if botoes[2].button("✅", key=f"{prefixo}_pagar_{indice}", help="Marcar como pago", type="secondary", use_container_width=True):
        try:
            df_atualizado = marcar_conta_como_paga(df_base, indice, usuario)
            if salvar_dados_empresa(normalizar_dataframe(df_atualizado), empresa) is None:
                return
            st.success("Conta marcada como paga.")
            st.rerun()
        except ValueError as erro:
            st.error(str(erro))
    if botoes[3].button("🗑️", key=f"{prefixo}_excluir_{indice}", help="Excluir conta", type="secondary", use_container_width=True):
        try:
            df_atualizado = excluir_conta_a_pagar(df_base, indice, usuario)
            if salvar_dados_empresa(normalizar_dataframe(df_atualizado), empresa) is None:
                return
            st.success("Conta excluída da lista ativa.")
            st.rerun()
        except ValueError as erro:
            st.error(str(erro))


def exibir_contas_em_cards(contas_exibidas: pd.DataFrame, df_base: pd.DataFrame, empresa: str, usuario: str) -> None:
    if contas_exibidas.empty:
        st.info("Nenhum registro encontrado para este filtro.")
        return

    controles = st.columns([1.8, 1.15, 1.15], gap="small")
    opcoes_ordenacao = list(ORDENACOES_CONTAS.keys())
    indice_atual = opcoes_ordenacao.index(st.session_state.get("contas_ordem_campo", "vencimento")) if st.session_state.get("contas_ordem_campo", "vencimento") in opcoes_ordenacao else 0
    campo = controles[0].selectbox(
        "Ordenar por",
        opcoes_ordenacao,
        index=indice_atual,
        format_func=lambda chave: ORDENACOES_CONTAS[chave]["label"],
        key="contas_cards_ordenar_por",
        label_visibility="visible",
    )
    direcao = controles[1].selectbox(
        "Direção",
        ["Decrescente", "Crescente"],
        index=0 if not bool(st.session_state.get("contas_ordem_crescente", True)) else 1,
        key="contas_cards_direcao",
        label_visibility="visible",
    )
    controles[2].caption("Use Cards em telas estreitas.")

    st.session_state.contas_ordem_campo = campo
    st.session_state.contas_ordem_crescente = direcao == "Crescente"

    contas = ordenar_contas_exibidas(contas_exibidas)
    grade = st.columns(2, gap="small")
    for posicao, (indice, linha) in enumerate(contas.iterrows()):
        nivel = nivel_prazo_conta(linha)
        indicador = indicador_prazo_conta(linha)
        alvo = grade[posicao % 2]
        with alvo:
            st.markdown(
                f"""
                <div class="conta-card">
                    <div class="conta-card-topo">
                        <div>
                            <div class="conta-card-label">Documento</div>
                            <div class="conta-card-documento">{escape(str(linha.get("documento", "")))}</div>
                        </div>
                        <div class="conta-card-status conta-status-{escape(nivel)}">{escape(str(linha.get("status", "")) if nivel != "vencida" else "vencido")}</div>
                    </div>
                    <div class="conta-card-grid">
                        <div class="conta-card-item">
                            <span class="conta-card-label">Vencimento</span>
                            <span class="conta-card-value">{escape(formatar_data_br(linha.get("vencimento")))}</span>
                        </div>
                        <div class="conta-card-item">
                            <span class="conta-card-label">Descrição</span>
                            <span class="conta-card-value">{escape(str(linha.get("descricao", "")))}</span>
                        </div>
                        <div class="conta-card-item">
                            <span class="conta-card-label">Fornecedor</span>
                            <span class="conta-card-value">{escape(str(linha.get("fornecedor_cliente", "")))}</span>
                        </div>
                        <div class="conta-card-item">
                            <span class="conta-card-label">Valor</span>
                            <span class="conta-card-value conta-card-valor">{escape(formatar_moeda_br(float(linha.get("valor", 0) or 0)))}</span>
                        </div>
                        <div class="conta-card-item">
                            <span class="conta-card-label">Status</span>
                            <span class="conta-card-value">{escape("vencido" if nivel == "vencida" else str(linha.get("status", "")))}</span>
                        </div>
                        <div class="conta-card-item">
                            <span class="conta-card-label">Indicador</span>
                            <span class="conta-card-value">{escape(indicador)}</span>
                        </div>
                    </div>
                </div>
                """,
                unsafe_allow_html=True,
            )
            _renderizar_acoes_conta(st, indice, df_base, empresa, usuario, "contas_card")


def limpar_nome_arquivo(nome: str) -> str:
    base = Path(nome).name
    return re.sub(r"[^A-Za-z0-9_.-]+", "_", base).strip("._") or "anexo"


def salvar_anexo(uploaded_file: object, empresa: str, documento: str) -> tuple[str, str]:
    if uploaded_file is None:
        return "", ""

    ANEXOS_DIR.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
    empresa_slug = limpar_nome_arquivo(empresa).upper()
    doc_slug = limpar_nome_arquivo(documento or "documento")
    nome_final = f"{timestamp}_{empresa_slug}_{doc_slug}_{limpar_nome_arquivo(uploaded_file.name)}"
    caminho = ANEXOS_DIR / nome_final
    caminho.write_bytes(uploaded_file.getbuffer())
    return uploaded_file.name, str(caminho.relative_to(BASE_DIR))


def obter_tipo_conta(selecao: str) -> tuple[str, str]:
    texto = str(selecao or "").strip()
    if " - " not in texto:
        return "", texto

    codigo, nome = texto.split(" - ", 1)
    codigo = codigo.strip()
    return codigo, nome.strip() or TIPOS_CONTA_PAGAR.get(codigo, "")


def opcoes_tipo_conta_com_historico(df: pd.DataFrame) -> list[str]:
    opcoes = []
    if {"tipo_conta_codigo", "tipo_conta_nome"}.issubset(df.columns):
        for _, linha in df[["tipo_conta_codigo", "tipo_conta_nome"]].dropna(how="all").iterrows():
            codigo = str(linha.get("tipo_conta_codigo", "") or "").strip()
            nome = str(linha.get("tipo_conta_nome", "") or "").strip()
            if codigo and nome:
                opcoes.append(f"{codigo} - {nome}")
            elif nome:
                opcoes.append(nome)
            elif codigo:
                opcoes.append(codigo)

    opcoes.extend(f"{codigo} - {nome}" for codigo, nome in TIPOS_CONTA_PAGAR.items())
    unicas = []
    chaves_vistas = set()
    for opcao in opcoes:
        texto = str(opcao or "").strip()
        if not texto:
            continue
        chave = unicodedata.normalize("NFKD", texto).encode("ascii", "ignore").decode("ascii").casefold()
        chave = " ".join(chave.split())
        if chave in OPCOES_OCULTAS:
            continue
        if chave in chaves_vistas:
            continue
        chaves_vistas.add(chave)
        unicas.append(texto)
    return unicas


def adicionar_conta_a_pagar(
    df: pd.DataFrame,
    empresa: str,
    usuario: str,
    dados_formulario: dict,
) -> pd.DataFrame:
    anexo_nome, anexo_caminho = salvar_anexo(
        dados_formulario["anexo"],
        empresa,
        dados_formulario["documento"],
    )
    tipo_codigo, tipo_nome = obter_tipo_conta(dados_formulario["tipo_conta"])
    nova_linha = {
        "empresa": empresa,
        "competencia": "",
        "tipo": "conta_a_pagar",
        "descricao": dados_formulario["descricao"],
        "fornecedor_cliente": dados_formulario["fornecedor"],
        "vencimento": dados_formulario["vencimento"].strftime("%Y-%m-%d"),
        "pagamento_recebimento": "",
        "valor": float(dados_formulario["valor"]),
        "status": dados_formulario["status"],
        "categoria": dados_formulario["categoria"],
        "observacao": dados_formulario["observacao"],
        "documento": dados_formulario["documento"],
        "tipo_conta_codigo": tipo_codigo,
        "tipo_conta_nome": tipo_nome,
        "anexo_nome": anexo_nome,
        "anexo_caminho": anexo_caminho,
        "codigo_pagamento": dados_formulario["codigo_pagamento"],
        "criado_em": agora_br(),
        "criado_por": usuario,
        "excluido_em": "",
        "excluido_por": "",
        "ativo": True,
    }
    return pd.concat([df, pd.DataFrame([nova_linha])], ignore_index=True)


def excluir_conta_a_pagar(df: pd.DataFrame, indice: int, usuario: str) -> pd.DataFrame:
    df_atualizado = df.copy()
    if indice not in df_atualizado.index:
        raise ValueError("Conta selecionada não foi encontrada na base atual.")

    posicao = df_atualizado.index.get_loc(indice)
    if isinstance(posicao, slice):
        posicao = posicao.start
    elif not isinstance(posicao, int):
        posicoes = list(posicao)
        if not posicoes:
            raise ValueError("Conta selecionada não foi encontrada na base atual.")
        posicao = posicoes[0]

    df_atualizado.iat[posicao, df_atualizado.columns.get_loc("ativo")] = False
    df_atualizado.iat[posicao, df_atualizado.columns.get_loc("excluido_em")] = agora_br()
    df_atualizado.iat[posicao, df_atualizado.columns.get_loc("excluido_por")] = usuario
    return df_atualizado


def marcar_conta_como_paga(df: pd.DataFrame, indice: int, usuario: str) -> pd.DataFrame:
    df_atualizado = df.copy()
    if indice not in df_atualizado.index:
        raise ValueError("Conta selecionada não foi encontrada na base atual.")

    posicao = df_atualizado.index.get_loc(indice)
    if isinstance(posicao, slice):
        posicao = posicao.start
    elif not isinstance(posicao, int):
        posicoes = list(posicao)
        if not posicoes:
            raise ValueError("Conta selecionada não foi encontrada na base atual.")
        posicao = posicoes[0]

    df_atualizado.iat[posicao, df_atualizado.columns.get_loc("status")] = "pago"
    df_atualizado.iat[posicao, df_atualizado.columns.get_loc("pagamento_recebimento")] = datetime.now().strftime("%Y-%m-%d")
    df_atualizado.iat[posicao, df_atualizado.columns.get_loc("observacao")] = (
        str(df_atualizado.iat[posicao, df_atualizado.columns.get_loc("observacao")] or "").strip()
        + f" | Pago registrado por {usuario} em {agora_br()}"
    ).strip(" |")
    return df_atualizado


def editar_conta_a_pagar(df: pd.DataFrame, indice: int, usuario: str, dados_formulario: dict) -> pd.DataFrame:
    df_atualizado = df.copy()
    if indice not in df_atualizado.index:
        raise ValueError("Conta selecionada não foi encontrada na base atual.")

    posicao = df_atualizado.index.get_loc(indice)
    if isinstance(posicao, slice):
        posicao = posicao.start
    elif not isinstance(posicao, int):
        posicoes = list(posicao)
        if not posicoes:
            raise ValueError("Conta selecionada não foi encontrada na base atual.")
        posicao = posicoes[0]

    tipo_codigo, tipo_nome = obter_tipo_conta(dados_formulario["tipo_conta"])
    anexo_nome, anexo_caminho = salvar_anexo(
        dados_formulario.get("anexo"),
        str(df_atualizado.iat[posicao, df_atualizado.columns.get_loc("empresa")] or ""),
        dados_formulario["documento"],
    )
    atualizacoes = {
        "descricao": dados_formulario["descricao"],
        "fornecedor_cliente": dados_formulario["fornecedor"],
        "vencimento": dados_formulario["vencimento"].strftime("%Y-%m-%d"),
        "valor": float(dados_formulario["valor"]),
        "status": dados_formulario["status"],
        "categoria": dados_formulario["categoria"],
        "observacao": dados_formulario["observacao"],
        "documento": dados_formulario["documento"],
        "tipo_conta_codigo": tipo_codigo,
        "tipo_conta_nome": tipo_nome,
        "codigo_pagamento": dados_formulario["codigo_pagamento"],
        "criado_por": str(df_atualizado.iat[posicao, df_atualizado.columns.get_loc("criado_por")] or usuario),
    }
    if anexo_nome:
        atualizacoes["anexo_nome"] = anexo_nome
        atualizacoes["anexo_caminho"] = anexo_caminho
    for coluna, valor in atualizacoes.items():
        df_atualizado.iat[posicao, df_atualizado.columns.get_loc(coluna)] = valor
    return df_atualizado


def duplicar_conta_a_pagar(df: pd.DataFrame, indice: int, usuario: str, dados_formulario: dict) -> pd.DataFrame:
    df_atualizado = df.copy()
    if indice not in df_atualizado.index:
        raise ValueError("Conta selecionada nao foi encontrada na base atual.")

    posicao = df_atualizado.index.get_loc(indice)
    if isinstance(posicao, slice):
        posicao = posicao.start
    elif not isinstance(posicao, int):
        posicoes = list(posicao)
        if not posicoes:
            raise ValueError("Conta selecionada nao foi encontrada na base atual.")
        posicao = posicoes[0]

    linha = df_atualizado.iloc[posicao].to_dict()
    anexo_nome, anexo_caminho = salvar_anexo(
        dados_formulario.get("anexo"),
        str(linha.get("empresa", "") or ""),
        dados_formulario["documento"],
    )
    tipo_codigo, tipo_nome = obter_tipo_conta(dados_formulario["tipo_conta"])
    nova_linha = dict(linha)
    nova_linha.update(
        {
            "competencia": "",
            "tipo": "conta_a_pagar",
            "descricao": dados_formulario["descricao"],
            "fornecedor_cliente": dados_formulario["fornecedor"],
            "vencimento": dados_formulario["vencimento"].strftime("%Y-%m-%d"),
            "pagamento_recebimento": "",
            "valor": float(dados_formulario["valor"]),
            "status": dados_formulario["status"],
            "categoria": dados_formulario["categoria"],
            "observacao": dados_formulario["observacao"],
            "documento": dados_formulario["documento"],
            "tipo_conta_codigo": tipo_codigo,
            "tipo_conta_nome": tipo_nome,
            "anexo_nome": anexo_nome or str(linha.get("anexo_nome", "") or ""),
            "anexo_caminho": anexo_caminho or str(linha.get("anexo_caminho", "") or ""),
            "codigo_pagamento": dados_formulario["codigo_pagamento"],
            "criado_em": agora_br(),
            "criado_por": usuario,
            "excluido_em": "",
            "excluido_por": "",
            "ativo": True,
        }
    )
    return pd.concat([df, pd.DataFrame([nova_linha], columns=COLUNAS_DADOS)], ignore_index=True)


def selecionar_conta_para_edicao(indice: int) -> None:
    st.session_state.conta_edicao_indice = indice


def selecionar_conta_para_duplicacao(indice: int) -> None:
    st.session_state.conta_duplicacao_indice = indice


def selecionar_manutencao(secao: str) -> None:
    st.session_state.manutencao_ativa = secao


def formulario_inclusao(df: pd.DataFrame, empresa: str, usuario: str) -> None:
    st.markdown("### + Incluir conta a pagar")
    st.caption("A inclusao fica registrada com data, hora e usuario logado.")

    opcoes_tipo = opcoes_com_novo(opcoes_tipo_conta_com_historico(df))
    opcoes_descricao = opcoes_com_novo(opcoes_com_historico(df, "descricao", SUGESTOES_DESCRICAO))
    opcoes_fornecedor = opcoes_com_novo(opcoes_com_historico(df, "fornecedor_cliente", SUGESTOES_FORNECEDOR))

    c1, c2 = st.columns(2)
    descricao_sel, descricao_nova = campo_opcao_nova(
        c1,
        "Descricao",
        opcoes_descricao,
        "Digite a nova descricao",
        "inclusao_descricao",
    )
    fornecedor_sel, fornecedor_novo = campo_opcao_nova(
        c2,
        "Fornecedor",
        opcoes_fornecedor,
        "Digite o novo fornecedor",
        "inclusao_fornecedor",
    )
    tipo_conta_sel, tipo_conta_novo = campo_opcao_nova(
        st,
        "Tipo de conta",
        opcoes_tipo,
        "Digite o novo tipo de conta",
        "inclusao_tipo_conta",
    )

    with st.form("form_conta_a_pagar", clear_on_submit=True):
        c3, c4, c5 = st.columns([1, 1, 1])
        vencimento = campo_data_ptbr(c3, "Vencimento", value=datetime.now().date(), key="form_conta_a_pagar_vencimento")
        valor_texto = c4.text_input("Valor", placeholder="Ex.: 1.000,00")
        status = c5.selectbox("Status", ["aberto", "pendente", "vencido"])

        observacao = st.text_area("Descricao / observacao livre", height=88)
        c6, c7 = st.columns([1, 1])
        anexo = c6.file_uploader("Anexo", type=["pdf", "png", "jpg", "jpeg", "xml", "csv", "xlsx"])
        codigo_pagamento = c7.text_area("Codigo de barras ou chave Pix", height=88)
        acao_cancelar, acao_salvar = st.columns([1, 2])
        cancelar = acao_cancelar.form_submit_button("Cancelar", use_container_width=True)
        enviar = acao_salvar.form_submit_button("Salvar conta a pagar", use_container_width=True)

    if cancelar:
        st.session_state.manutencao_ativa = ""
        st.rerun()

    if not enviar:
        return

    descricao = resolver_opcao_digitavel(descricao_sel, descricao_nova)
    fornecedor = resolver_opcao_digitavel(fornecedor_sel, fornecedor_novo)
    tipo_conta = resolver_opcao_digitavel(tipo_conta_sel, tipo_conta_novo)
    _, tipo_nome = obter_tipo_conta(tipo_conta)
    valor = parse_valor_br(valor_texto)

    if not descricao.strip() or not fornecedor.strip() or not tipo_conta.strip() or valor is None or valor <= 0:
        st.error("Preencha descricao, fornecedor, tipo de conta e valor maior que zero. Use ponto para milhar e virgula para centavos.")
        return
    if anexo is None and not codigo_pagamento.strip():
        st.error("Inclua um anexo ou informe o codigo de barras/chave Pix.")
        return

    documento_final = f"AP-{empresa}-{datetime.now().strftime('%Y%m%d%H%M%S')}"
    dados_formulario = {
        "descricao": descricao.strip(),
        "fornecedor": fornecedor.strip(),
        "vencimento": vencimento,
        "valor": valor,
        "status": status,
        "tipo_conta": tipo_conta,
        "categoria": tipo_nome.strip(),
        "documento": documento_final,
        "observacao": observacao.strip(),
        "anexo": anexo,
        "codigo_pagamento": codigo_pagamento.strip(),
    }
    df_atualizado = adicionar_conta_a_pagar(df, empresa, usuario, dados_formulario)
    if salvar_dados_empresa(df_atualizado, empresa) is None:
        return
    st.success("Conta a pagar incluida com sucesso.")
    st.rerun()

def importacao_massa_contas(df: pd.DataFrame, empresa: str, usuario: str) -> None:
    st.markdown("### Importar contas por CSV")
    st.caption("Baixe o modelo, preencha as linhas e importe o arquivo CSV nesta empresa.")
    st.download_button(
        "Baixar modelo CSV",
        data=gerar_modelo_importacao_csv(),
        file_name="modelo_contas_a_pagar.csv",
        mime="text/csv",
        use_container_width=True,
    )

    with st.form("form_importacao_massa", clear_on_submit=True):
        arquivo = st.file_uploader("Arquivo CSV preenchido", type=["csv"])
        importar = st.form_submit_button("Importar contas do CSV", use_container_width=True)

    if not importar:
        return

    if arquivo is None:
        st.error("Selecione um arquivo CSV para importar.")
        return

    try:
        df_importado = ler_csv_importacao(arquivo)
        novas_contas, erros = preparar_importacao_contas(df_importado, empresa, usuario)
    except ValueError as erro:
        st.error(str(erro))
        return

    if erros:
        st.warning("Algumas linhas não foram importadas.")
        for erro in erros[:10]:
            st.write(f"- {erro}")
        if len(erros) > 10:
            st.write(f"- Mais {len(erros) - 10} erro(s).")

    if novas_contas.empty:
        st.error("Nenhuma conta válida encontrada para importar.")
        return

    df_atualizado = pd.concat([df, novas_contas], ignore_index=True)
    if salvar_dados_empresa(df_atualizado, empresa) is None:
        return
    st.success(f"{len(novas_contas)} conta(s) importada(s) com sucesso.")
    st.rerun()


def formulario_edicao_conta(df: pd.DataFrame, indice: int, empresa: str, usuario: str, key_prefix: str) -> None:
    if indice not in df.index:
        st.error("Conta selecionada nao foi encontrada.")
        if st.button("Fechar", key=f"{key_prefix}_fechar_indice_invalido"):
            st.session_state.pop("conta_edicao_indice", None)
            st.rerun()
        return

    linha = df.loc[indice]
    vencimento_atual = pd.to_datetime(linha.get("vencimento"), errors="coerce")
    if pd.isna(vencimento_atual):
        vencimento_atual = pd.Timestamp(datetime.now().date())

    opcoes_tipo = opcoes_com_novo(opcoes_tipo_conta_com_historico(df))
    tipo_codigo_atual = str(linha.get("tipo_conta_codigo", "") or "").strip()
    tipo_nome_atual = str(linha.get("tipo_conta_nome", "") or "").strip()
    tipo_rotulo = f"{tipo_codigo_atual} - {tipo_nome_atual}" if tipo_codigo_atual and tipo_nome_atual else tipo_nome_atual
    if not tipo_rotulo:
        tipo_rotulo = opcoes_tipo[indice_padrao_sem_novo(opcoes_tipo)]
    if tipo_rotulo not in opcoes_tipo:
        opcoes_tipo.insert(0, tipo_rotulo)
    tipo_idx = opcoes_tipo.index(tipo_rotulo)

    st.caption(
        f"{formatar_data_br(linha.get('vencimento'))} | "
        f"{linha.get('fornecedor_cliente', '')} | "
        f"{formatar_moeda_br(float(linha.get('valor', 0) or 0))}"
    )

    tipo_conta_sel, tipo_conta_novo = campo_opcao_nova(
        st,
        "Tipo de conta",
        opcoes_tipo,
        "Digite o novo tipo de conta",
        f"{key_prefix}_tipo_conta",
        index=tipo_idx,
    )

    with st.form(f"{key_prefix}_form_editar_conta"):
        c1, c2 = st.columns(2)
        descricao = c1.text_input("Descricao", value=str(linha.get("descricao", "")))
        fornecedor = c2.text_input("Fornecedor", value=str(linha.get("fornecedor_cliente", "")))

        c3, c4, c5 = st.columns([1, 1, 1])
        vencimento = campo_data_ptbr(
            c3,
            "Vencimento",
            value=vencimento_atual.date(),
            key=f"{key_prefix}_vencimento",
        )
        valor_texto = c4.text_input(
            "Valor",
            value=formatar_moeda_br(float(linha.get("valor", 0) or 0)).replace("R$ ", ""),
            placeholder="Ex.: 1.000,00",
        )
        status_opcoes = ["aberto", "pendente", "vencido"]
        status_atual = str(linha.get("status", "aberto")).lower()
        status = c5.selectbox("Status", status_opcoes, index=status_opcoes.index(status_atual) if status_atual in status_opcoes else 0)

        observacao = st.text_area(
            "Descricao / observacao livre",
            value=str(linha.get("observacao", "") or "").strip(),
            height=88,
        )
        anexo_atual = str(linha.get("anexo_nome", "") or "").strip()
        if anexo_atual:
            st.caption(f"Anexo atual: {anexo_atual}")
        c6, c7 = st.columns([1, 1])
        anexo = c6.file_uploader("Novo anexo", type=["pdf", "png", "jpg", "jpeg", "xml", "csv", "xlsx"])
        codigo_pagamento = c7.text_area(
            "Codigo de barras ou chave Pix",
            value=str(linha.get("codigo_pagamento", "") or "").strip(),
            height=88,
        )
        b1, b2 = st.columns(2)
        salvar = b1.form_submit_button("Salvar alteracoes", use_container_width=True)
        cancelar = b2.form_submit_button("Cancelar", use_container_width=True)

    if cancelar:
        st.session_state.pop("conta_edicao_indice", None)
        st.rerun()
    if not salvar:
        return

    tipo_conta = resolver_opcao_digitavel(tipo_conta_sel, tipo_conta_novo)
    _, tipo_nome = obter_tipo_conta(tipo_conta)
    valor = parse_valor_br(valor_texto)
    if not descricao.strip() or not fornecedor.strip() or not tipo_conta.strip() or valor is None or valor <= 0:
        st.error("Preencha descricao, fornecedor, tipo de conta e valor maior que zero. Use ponto para milhar e virgula para centavos.")
        return
    if not str(linha.get("anexo_nome", "") or "").strip() and anexo is None and not codigo_pagamento.strip():
        st.error("Inclua um anexo ou informe o codigo de barras/chave Pix.")
        return

    dados_formulario = {
        "descricao": descricao.strip(),
        "fornecedor": fornecedor.strip(),
        "vencimento": vencimento,
        "valor": valor,
        "status": status,
        "tipo_conta": tipo_conta,
        "categoria": tipo_nome.strip(),
        "documento": str(linha.get("documento", "") or "").strip() or f"AP-{empresa}-{datetime.now().strftime('%Y%m%d%H%M%S')}",
        "observacao": observacao.strip(),
        "anexo": anexo,
        "codigo_pagamento": codigo_pagamento.strip(),
    }
    try:
        df_atualizado = editar_conta_a_pagar(df, indice, usuario, dados_formulario)
    except ValueError as erro:
        st.error(str(erro))
        return

    if salvar_dados_empresa(normalizar_dataframe(df_atualizado), empresa) is None:
        return
    st.session_state.pop("conta_edicao_indice", None)
    st.success("Conta a pagar atualizada com sucesso.")
    st.rerun()


def formulario_duplicacao_conta(df: pd.DataFrame, indice: int, empresa: str, usuario: str, key_prefix: str) -> None:
    if indice not in df.index:
        st.error("Conta selecionada nao foi encontrada.")
        if st.button("Fechar", key=f"{key_prefix}_fechar_indice_invalido"):
            st.session_state.pop("conta_duplicacao_indice", None)
            st.session_state.manutencao_ativa = ""
            st.rerun()
        return

    linha = df.loc[indice]
    vencimento_atual = pd.to_datetime(linha.get("vencimento"), errors="coerce")
    if pd.isna(vencimento_atual):
        vencimento_atual = pd.Timestamp(datetime.now().date())

    opcoes_tipo = opcoes_com_novo(opcoes_tipo_conta_com_historico(df))
    tipo_codigo_atual = str(linha.get("tipo_conta_codigo", "") or "").strip()
    tipo_nome_atual = str(linha.get("tipo_conta_nome", "") or "").strip()
    tipo_rotulo = f"{tipo_codigo_atual} - {tipo_nome_atual}" if tipo_codigo_atual and tipo_nome_atual else tipo_nome_atual
    if not tipo_rotulo:
        tipo_rotulo = opcoes_tipo[indice_padrao_sem_novo(opcoes_tipo)]
    if tipo_rotulo not in opcoes_tipo:
        opcoes_tipo.insert(0, tipo_rotulo)
    tipo_idx = opcoes_tipo.index(tipo_rotulo)

    st.caption(
        "Ao duplicar, ajuste a data do tributo e o valor antes de salvar. Os demais campos permanecem iguais e podem ser editados."
    )
    st.caption(
        f"{formatar_data_br(linha.get('vencimento'))} | "
        f"{linha.get('fornecedor_cliente', '')} | "
        f"{formatar_moeda_br(float(linha.get('valor', 0) or 0))}"
    )

    tipo_conta_sel, tipo_conta_novo = campo_opcao_nova(
        st,
        "Tipo de conta",
        opcoes_tipo,
        "Digite o novo tipo de conta",
        f"{key_prefix}_tipo_conta",
        index=tipo_idx,
    )

    with st.form(f"{key_prefix}_form_duplicar_conta"):
        c1, c2 = st.columns(2)
        descricao = c1.text_input("Descricao", value=str(linha.get("descricao", "")))
        fornecedor = c2.text_input("Fornecedor", value=str(linha.get("fornecedor_cliente", "")))

        c3, c4, c5 = st.columns([1, 1, 1])
        vencimento = campo_data_ptbr(
            c3,
            "Data do tributo",
            value=None,
            key=f"{key_prefix}_data_tributo",
            placeholder=vencimento_atual.strftime("%d/%m/%Y"),
        )
        valor_texto = c4.text_input("Valor do tributo", value="", placeholder="Ex.: 1.000,00")
        status_opcoes = ["aberto", "pendente", "vencido"]
        status_atual = str(linha.get("status", "aberto")).lower()
        status = c5.selectbox("Status", status_opcoes, index=status_opcoes.index(status_atual) if status_atual in status_opcoes else 0)

        observacao = st.text_area(
            "Descricao / observacao livre",
            value=str(linha.get("observacao", "") or "").strip(),
            height=88,
        )
        anexo_atual = str(linha.get("anexo_nome", "") or "").strip()
        if anexo_atual:
            st.caption(f"Anexo atual: {anexo_atual}")
        c6, c7 = st.columns([1, 1])
        anexo = c6.file_uploader("Novo anexo", type=["pdf", "png", "jpg", "jpeg", "xml", "csv", "xlsx"])
        codigo_pagamento = c7.text_area(
            "Codigo de barras ou chave Pix",
            value=str(linha.get("codigo_pagamento", "") or "").strip(),
            height=88,
        )
        b1, b2 = st.columns(2)
        salvar = b1.form_submit_button("Salvar duplicacao", use_container_width=True)
        cancelar = b2.form_submit_button("Cancelar", use_container_width=True)

    if cancelar:
        st.session_state.pop("conta_duplicacao_indice", None)
        st.session_state.manutencao_ativa = ""
        st.rerun()
    if not salvar:
        return

    tipo_conta = resolver_opcao_digitavel(tipo_conta_sel, tipo_conta_novo)
    _, tipo_nome = obter_tipo_conta(tipo_conta)
    valor = parse_valor_br(valor_texto)
    if vencimento is None or valor is None or valor <= 0:
        st.error("Informe uma data de tributo valida e um valor maior que zero antes de salvar a duplicacao.")
        return
    if not descricao.strip() or not fornecedor.strip() or not tipo_conta.strip():
        st.error("Preencha descricao, fornecedor e tipo de conta.")
        return
    if not str(linha.get("anexo_nome", "") or "").strip() and anexo is None and not codigo_pagamento.strip():
        st.error("Inclua um anexo ou informe o codigo de barras/chave Pix.")
        return

    documento_base = str(linha.get("documento", "") or "").strip()
    documento_novo = documento_base or f"AP-{empresa}-{datetime.now().strftime('%Y%m%d%H%M%S')}"
    dados_formulario = {
        "descricao": descricao.strip(),
        "fornecedor": fornecedor.strip(),
        "vencimento": vencimento.to_pydatetime().date(),
        "valor": valor,
        "status": status,
        "tipo_conta": tipo_conta,
        "categoria": tipo_nome.strip(),
        "documento": documento_novo,
        "observacao": observacao.strip(),
        "anexo": anexo,
        "codigo_pagamento": codigo_pagamento.strip(),
    }
    try:
        df_atualizado = duplicar_conta_a_pagar(df, indice, usuario, dados_formulario)
    except ValueError as erro:
        st.error(str(erro))
        return

    if salvar_dados_empresa(normalizar_dataframe(df_atualizado), empresa) is None:
        return
    st.session_state.pop("conta_duplicacao_indice", None)
    st.session_state.manutencao_ativa = ""
    st.success("Conta duplicada com sucesso.")
    st.rerun()


if hasattr(st, "dialog"):
    @st.dialog("Editar conta a pagar")
    def janela_edicao_conta(df: pd.DataFrame, indice: int, empresa: str, usuario: str) -> None:
        formulario_edicao_conta(df, indice, empresa, usuario, "janela_edicao")

    @st.dialog("Duplicar conta a pagar")
    def janela_duplicacao_conta(df: pd.DataFrame, indice: int, empresa: str, usuario: str) -> None:
        formulario_duplicacao_conta(df, indice, empresa, usuario, "janela_duplicacao")
else:
    def janela_edicao_conta(df: pd.DataFrame, indice: int, empresa: str, usuario: str) -> None:
        st.warning("Atualize o Streamlit para abrir edicao em janela. Usando edicao na pagina.")
        formulario_edicao_conta(df, indice, empresa, usuario, "janela_edicao")

    def janela_duplicacao_conta(df: pd.DataFrame, indice: int, empresa: str, usuario: str) -> None:
        st.warning("Atualize o Streamlit para abrir duplicacao em janela. Usando duplicacao na pagina.")
        formulario_duplicacao_conta(df, indice, empresa, usuario, "janela_duplicacao")


def area_edicao(df: pd.DataFrame, contas: pd.DataFrame, empresa: str, usuario: str) -> None:
    st.markdown("### Editar conta a pagar")
    st.caption("Selecione uma conta ativa, altere os campos e salve.")

    if contas.empty:
        st.info("Nenhuma conta em aberto para editar.")
        return

    opcoes = {}
    for indice, linha in contas.iterrows():
        rotulo = (
            f"{linha.get('documento', '')} | {formatar_data_br(linha.get('vencimento'))} | "
            f"{linha.get('fornecedor_cliente', '')} | {formatar_moeda_br(linha.get('valor', 0))}"
        )
        opcoes[rotulo] = indice

    labels = list(opcoes.keys())
    selecionada = st.selectbox("Conta a editar", labels, key="editar_conta_select")
    formulario_edicao_conta(df, opcoes[selecionada], empresa, usuario, "area_edicao")


def area_exclusao(df: pd.DataFrame, contas: pd.DataFrame, empresa: str, usuario: str) -> None:
    st.markdown("### 🗑️ Excluir conta a pagar")
    st.caption("A exclusão mantém auditoria no arquivo, marcando o registro como inativo.")

    if contas.empty:
        st.info("Nenhuma conta em aberto para excluir neste filtro.")
        return

    opcoes = {}
    for indice, linha in contas.iterrows():
        rotulo = (
            f"{linha.get('documento', '')} | {formatar_data_br(linha.get('vencimento'))} | "
            f"{linha.get('fornecedor_cliente', '')} | {formatar_moeda_br(linha.get('valor', 0))}"
        )
        opcoes[rotulo] = indice

    with st.form("form_excluir_conta"):
        selecionada = st.selectbox("Conta a excluir", list(opcoes.keys()))
        confirmar = st.checkbox("Confirmo a exclusão desta conta a pagar")
        excluir = st.form_submit_button("Excluir conta selecionada", use_container_width=True)

    if excluir:
        if not confirmar:
            st.warning("Marque a confirmação antes de excluir.")
            return

        df_atualizado = excluir_conta_a_pagar(df, opcoes[selecionada], usuario)
        if salvar_dados_empresa(normalizar_dataframe(df_atualizado), empresa) is None:
            return
        st.success("Conta a pagar excluída da lista ativa.")
        st.rerun()


def pagina_contas_a_pagar(df: pd.DataFrame, empresa: str, usuario: str) -> None:
    resumo_base = obter_resumo_base_empresa(empresa, df)
    contas = filtrar_contas_a_pagar_abertas(df, empresa)
    hoje = pd.Timestamp(datetime.now().date())
    total_aberto = contas["valor"].sum()
    vencidas = contas.loc[contas.apply(conta_vencida, axis=1)]
    vence_hoje = contas.loc[contas["vencimento_dt"].dt.date == hoje.date()]
    proximos_7 = contas.loc[
        (contas["vencimento_dt"] > hoje)
        & (contas["vencimento_dt"] <= hoje + pd.Timedelta(days=7))
    ]
    abertas_no_prazo = contas.loc[
        (contas["status"].astype(str).str.lower().isin(["aberto", "pendente"]))
        & (contas["vencimento_dt"] > hoje + pd.Timedelta(days=7))
    ]

    st.subheader("Painel de pagamentos")
    st.caption("Visao para acompanhar o que precisa ser pago por vencimento.")

    renderizar_metricas_em_duas_linhas(
        [
            {"label": "Total a pagar", "value": formatar_moeda_br(total_aberto), "tone": "neutral", "emoji": "💼"},
            {"label": "Vencidas", "value": formatar_moeda_br(vencidas["valor"].sum()), "delta": int(len(vencidas)), "tone": "danger", "emoji": "🔴"},
            {"label": "Vence hoje", "value": formatar_moeda_br(vence_hoje["valor"].sum()), "delta": int(len(vence_hoje)), "tone": "warning", "emoji": "🟠"},
            {"label": "Proximos 7 dias", "value": formatar_moeda_br(proximos_7["valor"].sum()), "delta": int(len(proximos_7)), "tone": "info", "emoji": "🟢"},
            {"label": "Contas em aberto", "value": int(len(contas)), "tone": "neutral", "emoji": "📄"},
            {"label": "No prazo", "value": int(len(abertas_no_prazo)), "tone": "ok", "emoji": "✅"},
            {
                "label": "Ticket medio",
                "value": formatar_moeda_br(total_aberto / len(contas)) if len(contas) else formatar_moeda_br(0),
                "tone": "warning",
                "emoji": "📊",
            },
        ]
    )

    if not len(contas):
        if resumo_base.get("status") == "ok":
            st.success("Nenhuma conta a pagar em aberto para esta empresa.")

    cabecalho_contas, acao_excel, visibilidade = st.columns([4.9, 1.15, 1.35], gap="small")
    with cabecalho_contas:
        st.markdown("### Contas para pagar")
    with acao_excel:
        st.download_button(
            "?? Excel",
            data=gerar_excel_contas_download(contas, empresa),
            file_name=f"contas_a_pagar_{slug_empresa(empresa)}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="baixar_excel_contas",
            use_container_width=False,
        )
    with visibilidade:
        st.selectbox(
            "Visualiza??o",
            ["Tabela", "Cards"],
            key="contas_visualizacao_modo",
            label_visibility="visible",
        )

    if contas.empty:
        if resumo_base.get("status") == "ok":
            st.info("Nenhum registro encontrado para este filtro.")
    else:
        if st.session_state.get("contas_visualizacao_modo", "Tabela") == "Cards":
            exibir_contas_em_cards(contas, df, empresa, usuario)
        else:
            exibir_contas_com_acoes(contas, df, empresa, usuario)

    indice_edicao = st.session_state.get("conta_edicao_indice")
    if indice_edicao is not None:
        janela_edicao_conta(df, indice_edicao, empresa, usuario)

    indice_duplicacao = st.session_state.get("conta_duplicacao_indice")
    if indice_duplicacao is not None:
        janela_duplicacao_conta(df, indice_duplicacao, empresa, usuario)

    st.divider()
    st.markdown("### Manutencao")

    st.session_state.setdefault("manutencao_ativa", "")
    a1, a2, a3, a4, espaco = st.columns([1.0, 1.0, 1.0, 1.0, 4.8], gap="small")
    a1.button("➕ Incluir", help="Incluir nova conta a pagar", key="btn_manutencao_incluir", on_click=selecionar_manutencao, args=("incluir",), use_container_width=True)
    a2.button("📥 Importar", help="Importar contas em massa por CSV", key="btn_manutencao_importar", on_click=selecionar_manutencao, args=("importar",), use_container_width=True)
    a3.button("🗑️ Excluir", help="Excluir conta a pagar", key="btn_manutencao_excluir", on_click=selecionar_manutencao, args=("excluir",), use_container_width=True)
    a4.button("📋 Duplicar", help="Duplicar conta a pagar", key="btn_manutencao_duplicar", on_click=selecionar_manutencao, args=("duplicar",), use_container_width=True)

    if st.session_state.manutencao_ativa == "incluir":
        st.markdown("#### ➕ Incluir nova conta a pagar")
        formulario_inclusao(df, empresa, usuario)
    elif st.session_state.manutencao_ativa == "importar":
        st.markdown("#### 📥 Importar contas em massa por CSV")
        importacao_massa_contas(df, empresa, usuario)
    elif st.session_state.manutencao_ativa == "excluir":
        st.markdown("#### 🗑️ Excluir conta a pagar")
        area_exclusao(df, contas, empresa, usuario)
    elif st.session_state.manutencao_ativa == "duplicar":
        st.markdown("#### 📋 Duplicar conta a pagar")
        indice_duplicacao = st.session_state.get("conta_duplicacao_indice")
        if indice_duplicacao is None:
            st.info("Selecione uma conta na tabela para duplicar.")
        else:
            formulario_duplicacao_conta(df, indice_duplicacao, empresa, usuario, "area_duplicacao")

    renderizar_salvamento_pendente()

    if resumo_base.get("status") in {"ausente", "vazia", "erro"}:
        st.divider()
        exibir_alerta_base_operacional(resumo_base)
        if resumo_base.get("status") == "vazia":
            st.warning(
                "A base operacional esta vazia. Se isso nao era esperado, restaure um backup antes de registrar novas contas."
            )

def main() -> None:
    configurar_pagina()
    aplicar_emojis_botoes_filtro()
    config = carregar_config()
    inicializar_sessao()
    restaurar_estado_sessao()
    expirar_sessao_se_necessario()

    if not st.session_state.autenticado:
        tela_login(config)
        st.stop()

    renderizar_menu_direito(config)

    if not st.session_state.get("empresa_logada"):
        dashboard_empresas(config, carregar_dados_empresas(config))
        st.stop()

    empresa, usuario = sidebar_filtros(config)
    df = carregar_dados_empresa(empresa)
    resumo_base = obter_resumo_base_empresa(empresa, df)
    contas = filtrar_contas_a_pagar_abertas(df, empresa)
    if resumo_base.get("status") == "ausente":
        status_geral, status_tipo = "Base operacional ausente", "alerta"
    elif resumo_base.get("status") == "vazia":
        status_geral, status_tipo = "Base operacional vazia", "alerta"
    elif resumo_base.get("status") == "erro":
        status_geral, status_tipo = "Falha ao ler base operacional", "alerta"
    else:
        status_geral, status_tipo = status_geral_contas(contas)

    mostrar_cabecalho(empresa, status_geral, status_tipo)
    if st.session_state.get("modulo_atual", "contas_a_pagar") == "contas_a_pagar":
        pagina_contas_a_pagar(df, empresa, usuario)
    else:
        st.info("Este módulo ainda não está liberado.")


if __name__ == "__main__":
    main()
